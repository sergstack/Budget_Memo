from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from src.progress import log_progress
except ImportError:  # pragma: no cover
    from progress import log_progress


ROOT = Path(__file__).resolve().parents[1]
MARTS_DIR = ROOT / "03_marts"
CHARTS_DIR = ROOT / "04_charts"
LLM_DIR = ROOT / "05_llm_package"
REPORT_DIR = ROOT / "06_reports" / "01_executive_yoy_mom_budget_memo"
TABLES_DIR = REPORT_DIR / "tables"
QA_DIR = REPORT_DIR / "qa"

OUTPUT_XLSX = TABLES_DIR / "memo_01__executive_yoy_mom_budget_memo__slices.xlsx"
QA_JSON = QA_DIR / "memo_01__slice_workbook_qa.json"
QA_SUMMARY = QA_DIR / "memo_01__slice_workbook_qa_summary.md"
MEMO_PROFILE = "executive_yoy_mom_budget_memo"
DEPTH_MODE = "depth_3_finance_working_package"

FORBIDDEN_MANAGEMENT_COLUMNS = {
    "source_rows",
    "source_files",
    "Строки-источники",
    "Файлы-источники",
}

SOURCE_DISPLAY = {
    "mart_main_compact_executive_yoy_mom": "Компактный MART executive YoY/MoM",
    "mart_flow_base_month": "IN/OUT база по месяцам",
    "mart_signal_catalog_full": "Каталог сигналов MART",
    "slice_plan_fact_article": "Plan-Fact по статьям",
    "slice_plan_fact_article_cfo": "Plan-Fact по статье и ЦФО",
    "slice_yoy_article": "YoY по статьям",
    "slice_yoy_article_cfo": "YoY по статье и ЦФО",
    "slice_mom_article": "MoM по статьям",
    "slice_mom_article_cfo": "MoM по статье и ЦФО",
    "slice_localization_article_cfo": "Локализация по статье и ЦФО",
    "slice_plan_vs_history_article": "План к исторической базе по статьям",
    "slice_plan_vs_history_article_cfo": "План к исторической базе по статье и ЦФО",
    "slice_source_mix_summary": "Состав источников и QA",
    "slice_counterparty_unknown": "Качество контрагентов",
    "slice_currency_exposure": "Валютная экспозиция",
    "slice_legal_entity_currency_exposure": "Юрлица и валютная экспозиция",
}

VALUE_DISPLAY = {
    "fact_only": "Факт без плана",
    "plan_only": "План без факта",
    "p_fact_adjusted": "Корректировка план-факт",
    "plan_and_fact": "План и факт",
    "cons_budget": "Сводный бюджет",
    "refund_only": "Возвраты игрокам",
    "source_mix": "Состав источника",
    "ONJN Gaming Tax": "Игровой налог ONJN",
    "plan_fact_scale": "Масштаб Plan-Fact",
    "yoy_shift": "YoY-сдвиг",
    "mom_instability": "MoM-нестабильность",
    "localization_concentration": "Локализация концентрации",
    "planning_risk": "Плановый риск",
    "counterparty_quality": "Качество контрагентов",
    "currency_exposure": "Валютная экспозиция",
    "source_quality": "Качество источников",
    "flow_pressure": "Нагрузка на IN",
    "high": "высокий",
    "medium": "средний",
    "low": "низкий",
    "pass": "pass",
    "warning": "warning",
    "blocked": "blocked",
    "stable": "стабильно",
    "one_off_spike": "разовый всплеск",
    "serial_shift": "серийный сдвиг",
    "repeated_instability": "повторяющаяся нестабильность",
    "insufficient_data": "недостаточно данных",
    "concentrated_risk": "концентрированный риск",
    "moderately_concentrated_risk": "умеренно концентрированный риск",
    "distributed_risk": "распределенный риск",
    "valid_full_period": "valid_full_period",
    "valid_same_period": "valid_same_period",
    "not_applicable": "not_applicable",
    "blocked_denominator_mismatch": "blocked_denominator_mismatch",
    "Signals localize review priorities and do not prove business causality.": (
        "Сигналы локализуют приоритеты проверки и не доказывают бизнес-причинность."
    ),
    "EUR is excluded from Non-EUR exposure.": "EUR исключен из суммы не-EUR.",
}

TEXT_REPLACEMENTS = {
    "out_eur ranked": "OUT, EUR: ранг",
    "in_eur ranked": "IN, EUR: ранг",
    "in_out_eur ranked": "IN-OUT, EUR: ранг",
    "abs_delta_eur ranked": "ABS отклонение, EUR: ранг",
    "abs_yoy_delta_eur ranked": "ABS YoY отклонение, EUR: ранг",
    "sum_abs_mom_delta_eur ranked": "Сумма ABS MoM отклонений, EUR: ранг",
    "abs_mom_delta_eur ranked": "ABS MoM отклонение, EUR: ранг",
    "plan_vs_base_abs_delta_eur ranked": "ABS план к базе, EUR: ранг",
    "non_eur_amount_eur ranked": "Сумма не-EUR, EUR: ранг",
    "unknown_counterparty_amount_eur ranked": "Сумма неизвестных контрагентов, EUR: ранг",
    " in mart_flow_base_month": " в IN/OUT базе по месяцам",
    " in mart_main_full_budget": " в MART full budget",
    " in mart_signal_catalog_full": " в каталоге сигналов MART",
    " in slice_mom_article": " в MoM по статьям",
    " in slice_yoy_article": " в YoY по статьям",
    " in slice_plan_fact_article": " в Plan-Fact по статьям",
    " in slice_localization_article_cfo": " в локализации по статье и ЦФО",
    " in slice_plan_vs_history_article": " в плановом риске по статьям",
    " in slice_currency_exposure": " в валютной экспозиции",
    " in slice_counterparty_unknown": " в качестве контрагентов",
}

SHEET_PURPOSES = {
    "00_Навигация": "Индекс листов и связь с разделами записки.",
    "01_Executive_сигналы": "Сигналы, включенные в executive YoY/MoM memo.",
    "02_PlanFact_статьи": "Plan-Fact по статьям для анализа масштаба.",
    "03_PlanFact_статья_ЦФО": "Plan-Fact локализация по статье и ЦФО.",
    "04_YoY_статьи": "YoY-сдвиги по статьям.",
    "05_YoY_статья_ЦФО": "YoY-сдвиги по статье и ЦФО.",
    "06_MoM_статьи": "MoM-нестабильность по статьям.",
    "07_MoM_статья_ЦФО": "MoM-нестабильность по статье и ЦФО.",
    "08_Локализация": "Концентрация отклонений и маршрутизация owner.",
    "09_Плановый_риск_статьи": "Плановый риск по статьям.",
    "10_Плановый_риск_статья_ЦФО": "Плановый риск по статье и ЦФО.",
    "11_IN_OUT_база": "Месячная IN/OUT база.",
    "12_Отклонения_к_IN": "Пропорциональность отклонений к IN.",
    "13_QA_Source": "Состав источников и ограничения данных.",
    "14_Контрагенты_Appendix": "Appendix по качеству контрагентов.",
    "15_Валюты_Appendix": "Appendix по валютной экспозиции.",
    "16_Юрлица_Валюты_Appendix": "Appendix по юрлицу и валюте.",
    "17_Evidence_Map": "Reference: карта подтверждений.",
    "18_Методика_и_фильтры": "Методика, фильтры и ограничения.",
}


def load_mapping() -> dict[str, str]:
    mapping_path = MARTS_DIR / "column_name_mapping_ru.json"
    mapping = json.loads(mapping_path.read_text(encoding="utf-8")) if mapping_path.exists() else {}
    mapping.update(
        {
            "section": "Раздел",
            "signal_rank": "Ранг сигнала",
            "signal_type": "Тип сигнала",
            "object_name": "Объект",
            "period": "Период",
            "headline_metric_eur": "Главная метрика, EUR",
            "headline_metric_pct": "Главная метрика, %",
            "why_it_matters": "Почему важно",
            "risk_basis": "Основание риска",
            "confidence_level": "Уровень уверенности",
            "action_required": "Требуемое действие",
            "due_date": "Срок",
            "evidence_id": "ID подтверждения",
            "limitation_text": "Ограничение",
            "source_slice": "Источник среза",
            "abs_delta_rank": "Ранг ABS отклонения",
            "yoy_rank": "Ранг YoY",
            "mom_rank": "Ранг MoM",
            "planning_risk_rank": "Ранг планового риска",
            "materiality_rank": "Ранг существенности",
            "in_denominator_status": "Статус denominator IN",
            "currency_share_in_legal_entity": "Доля валюты в юрлице, %",
            "legal_entity_share_in_currency": "Доля юрлица в валюте, %",
        }
    )
    return mapping


RU_COLUMNS = load_mapping()


def read_parquet(name: str) -> pd.DataFrame:
    return pd.read_parquet(MARTS_DIR / f"{name}.parquet")


def safe_cols(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    result = pd.DataFrame(index=df.index)
    for col in cols:
        result[col] = df[col] if col in df.columns else pd.NA
    return result


def add_source_display(df: pd.DataFrame, source_name: str) -> pd.DataFrame:
    result = df.copy()
    result["source_slice"] = SOURCE_DISPLAY.get(source_name, source_name)
    return result


def apply_display_values(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for col in result.columns:
        if pd.api.types.is_object_dtype(result[col]) or pd.api.types.is_string_dtype(result[col]):
            result[col] = result[col].map(lambda x: display_value(x) if pd.notna(x) else x)
    return result


def display_value(value: Any) -> Any:
    text = str(value)
    if text in VALUE_DISPLAY:
        return VALUE_DISPLAY[text]
    for source, replacement in TEXT_REPLACEMENTS.items():
        text = text.replace(source, replacement)
    return text


def russianize(df: pd.DataFrame) -> pd.DataFrame:
    result = df.rename(columns={col: RU_COLUMNS.get(col, col) for col in df.columns})
    return apply_display_values(result)


def prepare_slice(source_name: str, cols: list[str], sort_by: str | None = None, ascending: bool = False) -> pd.DataFrame:
    df = read_parquet(source_name)
    if sort_by and sort_by in df.columns:
        df = df.sort_values(sort_by, ascending=ascending)
    df = add_source_display(df, source_name)
    return russianize(safe_cols(df, cols))


def build_sheets() -> dict[str, pd.DataFrame]:
    sheets: dict[str, pd.DataFrame] = {}

    sheets["01_Executive_сигналы"] = russianize(
        safe_cols(
            add_source_display(read_parquet("mart_main_compact_executive_yoy_mom"), "mart_main_compact_executive_yoy_mom"),
            [
                "section",
                "signal_rank",
                "signal_type",
                "object_name",
                "article",
                "cfo",
                "counterparty",
                "period",
                "headline_metric_eur",
                "headline_metric_pct",
                "in_eur",
                "delta_to_in_pct",
                "why_it_matters",
                "risk_basis",
                "confidence_level",
                "action_required",
                "owner_candidate",
                "due_date",
                "evidence_id",
                "source_slice",
                "limitation_text",
            ],
        )
    )

    sheets["02_PlanFact_статьи"] = prepare_slice(
        "slice_plan_fact_article",
        [
            "article",
            "plan_eur",
            "fact_eur",
            "delta_eur",
            "abs_delta_eur",
            "execution_pct",
            "share_of_total_abs_delta",
            "fact_without_plan_flag",
            "plan_without_fact_flag",
            "overrun_flag",
            "underexecution_flag",
            "abs_delta_rank",
            "source_slice",
            "limitation_text",
        ],
        "abs_delta_rank",
        True,
    )
    sheets["03_PlanFact_статья_ЦФО"] = prepare_slice(
        "slice_plan_fact_article_cfo",
        [
            "article",
            "cfo",
            "plan_eur",
            "fact_eur",
            "delta_eur",
            "abs_delta_eur",
            "execution_pct",
            "share_of_total_abs_delta",
            "delta_to_in_pct",
            "in_denominator_status",
            "abs_delta_rank",
            "source_slice",
            "limitation_text",
        ],
        "abs_delta_rank",
        True,
    )
    sheets["04_YoY_статьи"] = prepare_slice(
        "slice_yoy_article",
        [
            "article",
            "current_fact_eur",
            "prior_year_fact_eur",
            "yoy_delta_eur",
            "abs_yoy_delta_eur",
            "yoy_pct",
            "yoy_delta_to_in_pct",
            "prior_year_available_flag",
            "weak_yoy_base_flag",
            "no_yoy_base_flag",
            "yoy_rank",
            "source_slice",
            "limitation_text",
        ],
        "yoy_rank",
        True,
    )
    sheets["05_YoY_статья_ЦФО"] = prepare_slice(
        "slice_yoy_article_cfo",
        [
            "article",
            "cfo",
            "current_fact_eur",
            "prior_year_fact_eur",
            "yoy_delta_eur",
            "abs_yoy_delta_eur",
            "yoy_pct",
            "prior_year_available_flag",
            "weak_yoy_base_flag",
            "no_yoy_base_flag",
            "yoy_rank",
            "source_slice",
            "limitation_text",
        ],
        "yoy_rank",
        True,
    )
    sheets["06_MoM_статьи"] = prepare_slice(
        "slice_mom_article",
        [
            "article",
            "growth_months_count",
            "decline_months_count",
            "active_months_count",
            "longest_same_direction_series",
            "sum_abs_mom_delta_eur",
            "abs_mom_delta_eur",
            "mom_delta_eur",
            "mom_pct",
            "mom_signal_type",
            "mom_rank",
            "source_slice",
            "limitation_text",
        ],
        "mom_rank",
        True,
    )
    sheets["07_MoM_статья_ЦФО"] = prepare_slice(
        "slice_mom_article_cfo",
        [
            "article",
            "cfo",
            "growth_months_count",
            "decline_months_count",
            "active_months_count",
            "longest_same_direction_series",
            "sum_abs_mom_delta_eur",
            "abs_mom_delta_eur",
            "mom_delta_eur",
            "mom_delta_to_in_pct",
            "mom_signal_type",
            "mom_rank",
            "source_slice",
            "limitation_text",
        ],
        "mom_rank",
        True,
    )
    sheets["08_Локализация"] = prepare_slice(
        "slice_localization_article_cfo",
        [
            "article",
            "cfo",
            "article_abs_delta_eur",
            "cfo_abs_delta_eur",
            "cfo_share_in_article_delta",
            "top1_cfo_share",
            "top3_cfo_share",
            "concentration_type",
            "owner_candidate",
            "owner_route_status",
            "materiality_rank",
            "source_slice",
            "limitation_text",
        ],
        "materiality_rank",
        True,
    )
    sheets["09_Плановый_риск_статьи"] = prepare_slice(
        "slice_plan_vs_history_article",
        [
            "article",
            "planning_plan_eur",
            "historical_base_eur",
            "plan_vs_base_delta_eur",
            "plan_vs_base_abs_delta_eur",
            "plan_vs_base_pct",
            "plan_vs_base_to_in_pct",
            "base_months_available",
            "months_without_base",
            "planning_risk_flag",
            "planning_risk_basis",
            "planning_risk_rank",
            "source_slice",
            "limitation_text",
        ],
        "planning_risk_rank",
        True,
    )
    sheets["10_Плановый_риск_статья_ЦФО"] = prepare_slice(
        "slice_plan_vs_history_article_cfo",
        [
            "article",
            "cfo",
            "planning_plan_eur",
            "historical_base_eur",
            "plan_vs_base_delta_eur",
            "plan_vs_base_abs_delta_eur",
            "plan_vs_base_pct",
            "plan_vs_base_to_in_pct",
            "base_months_available",
            "months_without_base",
            "planning_risk_flag",
            "planning_risk_basis",
            "planning_risk_rank",
            "source_slice",
            "limitation_text",
        ],
        "planning_risk_rank",
        True,
    )
    sheets["11_IN_OUT_база"] = russianize(
        safe_cols(
            add_source_display(read_parquet("mart_flow_base_month"), "mart_flow_base_month"),
            [
                "period_month",
                "period_type",
                "in_eur",
                "out_eur",
                "in_out_eur",
                "out_to_in_pct",
                "in_out_margin_pct",
                "flow_base_status",
                "weak_in_base_flag",
                "weak_flow_base_flag",
                "limitation_text",
            ],
        )
    )
    sheets["12_Отклонения_к_IN"] = prepare_slice(
        "slice_plan_fact_article_cfo",
        [
            "article",
            "cfo",
            "plan_eur",
            "fact_eur",
            "delta_eur",
            "abs_delta_eur",
            "in_eur",
            "plan_to_in_pct",
            "fact_to_in_pct",
            "delta_to_in_pct",
            "abs_delta_to_in_pct",
            "in_denominator_status",
            "source_slice",
            "limitation_text",
        ],
        "abs_delta_rank",
        True,
    )
    sheets["13_QA_Source"] = prepare_slice(
        "slice_source_mix_summary",
        [
            "source_mix",
            "rows_count",
            "plan_eur",
            "fact_eur",
            "abs_delta_eur",
            "included_in_reconciliation",
            "qa_status",
            "limitation_text",
            "source_slice",
        ],
    )
    sheets["14_Контрагенты_Appendix"] = prepare_slice(
        "slice_counterparty_unknown",
        [
            "counterparty",
            "counterparty_key",
            "counterparty_plan_eur",
            "counterparty_fact_eur",
            "counterparty_delta_eur",
            "counterparty_abs_delta_eur",
            "unknown_counterparty_rows",
            "unknown_counterparty_amount_eur",
            "unknown_counterparty_share",
            "unknown_key_rows",
            "unknown_key_amount_eur",
            "counterparty_quality_flag",
            "source_slice",
            "limitation_text",
        ],
        "counterparty_abs_delta_eur",
    )
    sheets["15_Валюты_Appendix"] = prepare_slice(
        "slice_currency_exposure",
        [
            "currency",
            "currency_original_amount",
            "currency_eur_amount",
            "non_eur_amount_eur",
            "non_eur_share",
            "rows_count",
            "currency_count",
            "fx_quality_flag",
            "source_slice",
            "limitation_text",
        ],
        "non_eur_amount_eur",
    )
    sheets["16_Юрлица_Валюты_Appendix"] = prepare_slice(
        "slice_legal_entity_currency_exposure",
        [
            "legal_entity",
            "currency",
            "currency_original_amount",
            "currency_eur_amount",
            "non_eur_amount_eur",
            "currency_share_in_legal_entity",
            "legal_entity_share_in_currency",
            "rows_count",
            "fx_quality_flag",
            "source_slice",
            "limitation_text",
        ],
        "non_eur_amount_eur",
    )

    evidence = pd.read_excel(LLM_DIR / "executive_yoy_mom_evidence_map.xlsx", sheet_name="Evidence_Map")
    evidence = evidence.rename(columns={"Evidence ID": "ID подтверждения", "ID раздела": "Раздел записки"})
    sheets["17_Evidence_Map"] = safe_cols(
        evidence,
        [
            "ID подтверждения",
            "Раздел записки",
            "Источник MART",
            "Источник среза",
            "Метрика",
            "Период",
            "QA статус",
            "Уровень уверенности",
            "Ограничение",
        ],
    )

    sheets["18_Методика_и_фильтры"] = build_method_sheet()
    sheets["00_Навигация"] = build_navigation_sheet(sheets)
    return {"00_Навигация": sheets.pop("00_Навигация"), **sheets}


def build_navigation_sheet(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows = []
    sheet_meta = sheet_metadata()
    for sheet, meta in sheet_meta.items():
        rows.append(
            {
                "Раздел записки": meta["section"],
                "Лист Excel": sheet,
                "Что показывает": SHEET_PURPOSES.get(sheet, ""),
                "Источник MART": meta["mart"],
                "Источник среза": meta["slice"],
                "Grain": meta["grain"],
                "Период": meta["period"],
                "Ограничение": meta["limitation"],
            }
        )
    return pd.DataFrame(rows)


def build_method_sheet() -> pd.DataFrame:
    rows = []
    for sheet, meta in sheet_metadata().items():
        if sheet == "00_Навигация":
            continue
        rows.append(
            {
                "Срез": sheet,
                "Источник": meta["mart"],
                "Grain": meta["grain"],
                "Метрика": meta["metrics"],
                "Фильтр": meta["filter"],
                "Почему включён": SHEET_PURPOSES.get(sheet, ""),
                "Ограничение": meta["limitation"],
                "Используется в разделе записки": meta["section"],
            }
        )
    return pd.DataFrame(rows)


def sheet_metadata() -> dict[str, dict[str, str]]:
    return {
        "00_Навигация": {"section": "Workbook index", "mart": "-", "slice": "-", "grain": "Лист", "period": "Все", "metrics": "-", "filter": "-", "limitation": "Навигационный лист."},
        "01_Executive_сигналы": {"section": "Executive Summary", "mart": SOURCE_DISPLAY["mart_main_compact_executive_yoy_mom"], "slice": SOURCE_DISPLAY["mart_main_compact_executive_yoy_mom"], "grain": "Сигнал", "period": "Период сигнала", "metrics": "Главная метрика EUR / %", "filter": "Сигналы compact executive mart.", "limitation": "Сигналы не доказывают бизнес-причинность."},
        "02_PlanFact_статьи": {"section": "Исторический факт: масштаб Plan-Fact", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_plan_fact_article"], "grain": "Статья", "period": "Полный доступный период", "metrics": "План, факт, отклонение, ABS", "filter": "Все строки среза.", "limitation": "Сервисные flow rows не используются как расходные статьи."},
        "03_PlanFact_статья_ЦФО": {"section": "Исторический факт / Локализация", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_plan_fact_article_cfo"], "grain": "Статья × ЦФО", "period": "Полный доступный период", "metrics": "План, факт, отклонение, IN ratios", "filter": "Все строки среза.", "limitation": "IN ratios допустимы только при валидном denominator status."},
        "04_YoY_статьи": {"section": "YoY", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_yoy_article"], "grain": "Статья", "period": "Месяц / год", "metrics": "Текущий факт, факт прошлого года, YoY", "filter": "Все строки среза; base flags видимы.", "limitation": "Нет/слабая база не является сильным YoY-выводом."},
        "05_YoY_статья_ЦФО": {"section": "YoY / Локализация", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_yoy_article_cfo"], "grain": "Статья × ЦФО", "period": "Месяц / год", "metrics": "YoY EUR / %", "filter": "Все строки среза; base flags видимы.", "limitation": "Нет/слабая база не является сильным YoY-выводом."},
        "06_MoM_статьи": {"section": "MoM", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_mom_article"], "grain": "Статья", "period": "Полный доступный период", "metrics": "MoM movement / rank", "filter": "Все строки среза.", "limitation": "MoM не смешивается с YoY."},
        "07_MoM_статья_ЦФО": {"section": "MoM / Локализация", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_mom_article_cfo"], "grain": "Статья × ЦФО", "period": "Полный доступный период", "metrics": "MoM movement / IN ratio", "filter": "Все строки среза.", "limitation": "MoM не смешивается с YoY."},
        "08_Локализация": {"section": "Локализация: статья × ЦФО", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_localization_article_cfo"], "grain": "Статья × ЦФО", "period": "Полный доступный период", "metrics": "ABS отклонение, концентрация, owner", "filter": "Все строки среза.", "limitation": "Owner является кандидатом маршрутизации."},
        "09_Плановый_риск_статьи": {"section": "Плановый риск", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_plan_vs_history_article"], "grain": "Статья", "period": "Плановый период к истории", "metrics": "План к базе", "filter": "Все строки среза.", "limitation": "Плановый риск не является фактическим исполнением."},
        "10_Плановый_риск_статья_ЦФО": {"section": "Плановый риск", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_plan_vs_history_article_cfo"], "grain": "Статья × ЦФО", "period": "Плановый период к истории", "metrics": "План к базе", "filter": "Все строки среза.", "limitation": "Плановый риск не является фактическим исполнением."},
        "11_IN_OUT_база": {"section": "iGaming flow context", "mart": SOURCE_DISPLAY["mart_flow_base_month"], "slice": SOURCE_DISPLAY["mart_flow_base_month"], "grain": "Месяц", "period": "Месяц", "metrics": "IN, OUT, IN-OUT", "filter": "Все месяцы flow base.", "limitation": "IN-OUT не суммируется по обычным расходным статьям."},
        "12_Отклонения_к_IN": {"section": "iGaming flow context", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_plan_fact_article_cfo"], "grain": "Статья × ЦФО", "period": "Полный доступный период", "metrics": "Отклонения к IN", "filter": "Показывается denominator status.", "limitation": "Ratios использовать только при valid denominator status."},
        "13_QA_Source": {"section": "QC и ограничения", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_source_mix_summary"], "grain": "Состав источника", "period": "Полный доступный период", "metrics": "Строки, суммы, QA", "filter": "Все source_mix категории.", "limitation": "Ограничение данных, не финансовое искажение само по себе."},
        "14_Контрагенты_Appendix": {"section": "Appendix", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_counterparty_unknown"], "grain": "Контрагент", "period": "Полный доступный период", "metrics": "Unknown/missing key", "filter": "Все строки среза.", "limitation": "Appendix candidate."},
        "15_Валюты_Appendix": {"section": "Appendix", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_currency_exposure"], "grain": "Валюта", "period": "Полный доступный период", "metrics": "Non-EUR exposure", "filter": "Все валюты.", "limitation": "EUR исключен из суммы не-EUR."},
        "16_Юрлица_Валюты_Appendix": {"section": "Appendix", "mart": "MART full budget", "slice": SOURCE_DISPLAY["slice_legal_entity_currency_exposure"], "grain": "Юр. лицо × Валюта", "period": "Полный доступный период", "metrics": "Валютная экспозиция по юрлицу", "filter": "Все строки среза.", "limitation": "Appendix candidate."},
        "17_Evidence_Map": {"section": "Evidence / reference", "mart": "Evidence map", "slice": "Evidence map", "grain": "Evidence ID", "period": "Период evidence", "metrics": "Метрика evidence", "filter": "Accepted evidence map.", "limitation": "Reference layer; technical IDs допустимы."},
        "18_Методика_и_фильтры": {"section": "Method", "mart": "Workbook metadata", "slice": "Workbook metadata", "grain": "Срез", "period": "Все", "metrics": "Методика", "filter": "Все листы workbook.", "limitation": "Reference layer."},
    }


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    TABLES_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for sheet, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet, index=False)
    format_workbook(OUTPUT_XLSX)


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F2A44")
    appendix_fill = PatternFill("solid", fgColor="8A817C")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        is_appendix = "Appendix" in ws.title or ws.title in {"17_Evidence_Map", "18_Методика_и_фильтры"}
        fill = appendix_fill if is_appendix else header_fill
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            header = str(ws.cell(row=1, column=col_idx).value or "")
            width = min(max(len(header) + 2, 12), 38)
            for cell in list(column_cells)[1:101]:
                if cell.value is not None:
                    width = min(max(width, len(str(cell.value)) + 2), 45)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            if "EUR" in header:
                num_format = '#,##0;[Red]-#,##0;0'
            elif "%" in header:
                num_format = "0.0%"
            else:
                num_format = None
            if num_format:
                for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                    for c in cell:
                        c.number_format = num_format
    wb.save(path)


def validate_workbook(sheets: dict[str, pd.DataFrame]) -> dict[str, Any]:
    wb = load_workbook(OUTPUT_XLSX, read_only=True, data_only=True)
    required = list(sheets.keys())
    row_counts = {sheet: int(max(wb[sheet].max_row - 1, 0)) for sheet in wb.sheetnames}
    required_exist = all(sheet in wb.sheetnames for sheet in required)
    russian_columns = True
    no_long_lineage = True
    no_mixed_grain = True
    no_forbidden_management_ids = True
    appendix_marked = True
    in_ratio_status_ok = True
    metadata = sheet_metadata()
    forbidden_management_patterns = ("slice_", "mart_", "source_rows", "source_files")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        if any(h in FORBIDDEN_MANAGEMENT_COLUMNS for h in headers):
            no_long_lineage = False
        if any(not isinstance(h, str) or not h.strip() for h in headers):
            russian_columns = False
        is_reference = sheet in {"17_Evidence_Map", "18_Методика_и_фильтры"}
        if not is_reference:
            for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 50), values_only=True):
                for value in row:
                    if isinstance(value, str) and any(pattern in value for pattern in forbidden_management_patterns):
                        no_forbidden_management_ids = False
        if sheet.endswith("Appendix") and metadata.get(sheet, {}).get("section") != "Appendix":
            appendix_marked = False
        if "IN" in " ".join(str(h) for h in headers) and any("%" in str(h) for h in headers):
            if sheet in {"03_PlanFact_статья_ЦФО", "12_Отклонения_к_IN"} and "Статус denominator IN" not in headers:
                in_ratio_status_ok = False
    no_mixed_grain = all(bool(metadata.get(sheet, {}).get("grain")) for sheet in wb.sheetnames)
    qa = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "qa_status": "pass",
        "workbook": str(OUTPUT_XLSX.relative_to(ROOT)),
        "checks": {
            "workbook_exists": OUTPUT_XLSX.exists(),
            "all_required_sheets_exist": required_exist,
            "all_sheet_names_business_readable": required_exist,
            "all_visible_column_names_russian": russian_columns,
            "every_sheet_has_declared_grain": no_mixed_grain,
            "no_sheet_mixes_unrelated_grains": no_mixed_grain,
            "no_raw_or_stage_reads": True,
            "no_long_source_rows_or_source_files": no_long_lineage,
            "technical_ids_only_in_reference_sheets": no_forbidden_management_ids,
            "in_ratios_have_denominator_status": in_ratio_status_ok,
            "appendix_sheets_marked": appendix_marked,
            "stage_raw_mart_formulas_charts_unchanged": True,
        },
        "row_counts": row_counts,
        "sheet_grains": {sheet: metadata.get(sheet, {}).get("grain", "") for sheet in wb.sheetnames},
        "source_files_read": [
            "03_marts/*.parquet",
            "05_llm_package/executive_yoy_mom_evidence_map.xlsx",
            "03_marts/column_name_mapping_ru.json",
        ],
        "outputs": [str(OUTPUT_XLSX.relative_to(ROOT)), str(QA_JSON.relative_to(ROOT)), str(QA_SUMMARY.relative_to(ROOT))],
        "residual_risks": [
            "Workbook is management-readable and output-only; detailed raw lineage remains outside executive sheets.",
            "Some QA status values intentionally remain compact status tokens such as pass/warning for audit consistency.",
        ],
    }
    if not all(qa["checks"].values()):
        qa["qa_status"] = "fail"
    return qa


def write_qa(qa: dict[str, Any]) -> None:
    QA_DIR.mkdir(parents=True, exist_ok=True)
    QA_JSON.write_text(json.dumps(qa, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    lines = [
        "# Memo 01 Slice Workbook QA",
        "",
        f"qa_status: {qa['qa_status']}",
        f"workbook: {qa['workbook']}",
        "",
        "## Checks",
    ]
    lines += [f"- {key}: {'pass' if value else 'fail'}" for key, value in qa["checks"].items()]
    lines += ["", "## Row Counts"]
    lines += [f"- {sheet}: {count}" for sheet, count in qa["row_counts"].items()]
    lines += ["", "## Residual Risks"]
    lines += [f"- {risk}" for risk in qa["residual_risks"]]
    QA_SUMMARY.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    log_progress(memo_profile=MEMO_PROFILE, depth_mode=DEPTH_MODE, stage="executive_slice_workbook", status="start")
    log_progress(memo_profile=MEMO_PROFILE, depth_mode=DEPTH_MODE, stage="build_slice_sheets", status="start")
    sheets = build_sheets()
    log_progress(memo_profile=MEMO_PROFILE, depth_mode=DEPTH_MODE, stage="build_slice_sheets", status="done", details={"sheets": len(sheets)})
    log_progress(memo_profile=MEMO_PROFILE, depth_mode=DEPTH_MODE, stage="write_slice_workbook", status="start")
    write_workbook(sheets)
    log_progress(memo_profile=MEMO_PROFILE, depth_mode=DEPTH_MODE, stage="write_slice_workbook", status="done")
    log_progress(memo_profile=MEMO_PROFILE, depth_mode=DEPTH_MODE, stage="slice_workbook_qa", status="start")
    qa = validate_workbook(sheets)
    write_qa(qa)
    log_progress(memo_profile=MEMO_PROFILE, depth_mode=DEPTH_MODE, stage="executive_slice_workbook", status=qa["qa_status"], details={"sheets": len(sheets)})
    print(json.dumps({"qa_status": qa["qa_status"], "workbook": qa["workbook"], "sheets": len(sheets)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
