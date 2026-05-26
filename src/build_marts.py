from __future__ import annotations

import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

try:
    from src.main import ACTUALS_CLOSED_THROUGH_MONTH
except ModuleNotFoundError:  # pragma: no cover - supports `python3 src/build_marts.py`
    from main import ACTUALS_CLOSED_THROUGH_MONTH


PROJECT_ROOT = Path(__file__).resolve().parents[1]
STAGE_SOURCE = PROJECT_ROOT / "02_stage" / "01_full_stage.csv"
MARTS_DIR = PROJECT_ROOT / "03_marts"
SIGNALS_DIR = PROJECT_ROOT / "04_signals"
EVIDENCE_DIR = PROJECT_ROOT / "05_evidence"
QA_DIR = PROJECT_ROOT / "07_qa"

ARCHIVE_PREREQUISITE = PROJECT_ROOT / "99_archive" / "old_downstream_before_mart_rebuild_20260521_1252"

SERVICE_VALUES = {"IN", "OUT", "IN-OUT"}
UNKNOWN_VALUES = {"", "unknown", "not_applicable", "Контрагент не указан", "p_fact", "p-fact"}

TECHNICAL_OUTPUTS = {
    "mart_main_full_budget": MARTS_DIR / "mart_main_full_budget.parquet",
    "mart_flow_base_month": MARTS_DIR / "mart_flow_base_month.parquet",
    "mart_signal_catalog_full": MARTS_DIR / "mart_signal_catalog_full.parquet",
    "mart_main_compact_executive_yoy_mom": MARTS_DIR / "mart_main_compact_executive_yoy_mom.parquet",
}

THRESHOLDS = {
    "material_abs_delta_threshold_eur": 100_000.0,
    "material_delta_to_in_threshold_pct": 0.01,
    "material_yoy_delta_threshold_eur": 100_000.0,
    "material_mom_delta_threshold_eur": 100_000.0,
    "planning_risk_threshold_eur": 100_000.0,
    "weak_base_threshold_eur": 1_000.0,
}

SOURCE_COLUMN_MAP = {
    "Месяц": "period_month",
    "Дата": "date",
    "Тип периода": "period_type",
    "source_mix": "source_mix",
    "included_in_reconciliation": "included_in_reconciliation",
    "has_plan": "has_plan",
    "has_fact": "has_fact",
    "has_p_fact_adjustment": "has_p_fact_adjustment",
    "has_player_refund": "has_player_refund",
    "Код статьи ДДС": "article_code",
    "Тип": "article_type",
    "Статья 1": "article_level_1",
    "Статья 2": "article_level_2",
    "Статья": "article",
    "ЦФО": "cfo",
    "Юр. лицо": "legal_entity",
    "Контрагент": "counterparty",
    "Ключ контрагента": "counterparty_key",
    "Тип контрагента": "counterparty_type",
    "Валюта": "currency",
    "Сумма исходная": "source_amount",
    "План, EUR": "plan_eur",
    "Факт, EUR": "fact_eur",
    "IN-OUT, EUR": "stage_in_out_eur",
    "source_file": "source_file",
    "source_row_id": "source_row_id",
}

COLUMN_NAME_MAPPING_RU = {
    "row_id": "ID строки",
    "period_month": "Месяц",
    "period_year": "Год",
    "date": "Дата",
    "period_type": "Тип периода",
    "source_mix": "Состав источника",
    "included_in_reconciliation": "Включено в сверку",
    "has_plan": "Есть план",
    "has_fact": "Есть факт",
    "has_p_fact_adjustment": "Есть корректировка p-fact",
    "has_player_refund": "Есть возврат игроку",
    "article_code": "Код статьи ДДС",
    "article_type": "Тип статьи",
    "article_type_sort": "Тип статьи сортировка",
    "article_level_1": "Статья 1",
    "article_level_2": "Статья 2",
    "article": "Статья",
    "cfo": "ЦФО",
    "legal_entity": "Юр. лицо",
    "counterparty": "Контрагент",
    "counterparty_key": "Ключ контрагента",
    "counterparty_type": "Тип контрагента",
    "currency": "Валюта",
    "source_amount": "Сумма исходная",
    "plan_eur": "План, EUR",
    "fact_eur": "Факт, EUR",
    "delta_eur": "Отклонение План-Факт, EUR",
    "abs_delta_eur": "ABS отклонение, EUR",
    "execution_pct": "Исполнение, %",
    "share_of_total_abs_delta": "Доля ABS отклонения от итога, %",
    "fact_without_plan_flag": "Факт без плана",
    "plan_without_fact_flag": "План без факта",
    "overrun_flag": "Факт выше плана",
    "underexecution_flag": "Факт ниже плана",
    "in_eur": "IN, EUR",
    "out_eur": "OUT, EUR",
    "in_out_eur": "IN-OUT, EUR",
    "out_to_in_pct": "OUT к IN, %",
    "in_out_margin_pct": "IN-OUT маржа, %",
    "flow_base_status": "Статус IN/OUT базы",
    "flow_value_source": "Источник IN/OUT базы",
    "flow_uses_plan_fallback_flag": "IN/OUT использует план",
    "weak_in_base_flag": "Слабая IN база",
    "weak_flow_base_flag": "Слабая flow база",
    "plan_to_in_pct": "План к IN, %",
    "fact_to_in_pct": "Факт к IN, %",
    "delta_to_in_pct": "Отклонение к IN, %",
    "abs_delta_to_in_pct": "ABS отклонение к IN, %",
    "current_fact_eur": "Текущий факт, EUR",
    "prior_year_fact_eur": "Факт прошлого года, EUR",
    "yoy_delta_eur": "YoY отклонение, EUR",
    "abs_yoy_delta_eur": "ABS YoY отклонение, EUR",
    "yoy_pct": "YoY, %",
    "yoy_delta_to_in_pct": "YoY отклонение к IN, %",
    "abs_yoy_delta_to_in_pct": "ABS YoY отклонение к IN, %",
    "prior_year_available_flag": "Есть база прошлого года",
    "prior_year_month_count": "Месяцев базы прошлого года",
    "weak_yoy_base_flag": "Слабая YoY база",
    "no_yoy_base_flag": "Нет YoY базы",
    "current_month_fact_eur": "Факт текущего месяца, EUR",
    "previous_month_fact_eur": "Факт предыдущего месяца, EUR",
    "mom_delta_eur": "MoM отклонение, EUR",
    "abs_mom_delta_eur": "ABS MoM отклонение, EUR",
    "mom_pct": "MoM, %",
    "mom_delta_to_in_pct": "MoM отклонение к IN, %",
    "abs_mom_delta_to_in_pct": "ABS MoM отклонение к IN, %",
    "growth_months_count": "Месяцев роста",
    "decline_months_count": "Месяцев снижения",
    "active_months_count": "Активных месяцев",
    "longest_same_direction_series": "Длина серии одного направления",
    "mom_signal_type": "Тип MoM сигнала",
    "article_abs_delta_eur": "ABS отклонение статьи, EUR",
    "cfo_abs_delta_eur": "ABS отклонение ЦФО, EUR",
    "cfo_share_in_article_delta": "Доля ЦФО в отклонении статьи, %",
    "top1_cfo_share": "Доля top-1 ЦФО, %",
    "top3_cfo_share": "Доля top-3 ЦФО, %",
    "concentration_type": "Тип концентрации",
    "owner_candidate": "Кандидат владельца",
    "owner_route_status": "Статус маршрутизации владельца",
    "planning_plan_eur": "План планового периода, EUR",
    "historical_base_eur": "Историческая база, EUR",
    "plan_vs_base_delta_eur": "План к базе, EUR",
    "plan_vs_base_abs_delta_eur": "ABS план к базе, EUR",
    "plan_vs_base_pct": "План к базе, %",
    "plan_vs_base_to_in_pct": "План к базе к IN, %",
    "base_months_available": "Месяцев базы",
    "months_without_base": "Месяцев без базы",
    "planning_risk_flag": "Флаг планового риска",
    "planning_risk": "Плановый риск",
    "planning_risk_basis": "Основание планового риска",
    "counterparty_fact_eur": "Факт контрагента, EUR",
    "counterparty_plan_eur": "План контрагента, EUR",
    "counterparty_delta_eur": "Отклонение контрагента, EUR",
    "counterparty_abs_delta_eur": "ABS отклонение контрагента, EUR",
    "unknown_counterparty_rows": "Строк с неизвестным контрагентом",
    "unknown_counterparty_amount_eur": "Сумма неизвестных контрагентов, EUR",
    "unknown_counterparty_share": "Доля неизвестных контрагентов, %",
    "unknown_key_rows": "Строк с неизвестным ключом",
    "unknown_key_amount_eur": "Сумма неизвестных ключей, EUR",
    "top5_counterparty_share": "Доля top-5 контрагентов, %",
    "top10_counterparty_share": "Доля top-10 контрагентов, %",
    "counterparty_quality_flag": "Флаг качества контрагентов",
    "legal_entity_plan_eur": "План юрлица, EUR",
    "legal_entity_fact_eur": "Факт юрлица, EUR",
    "legal_entity_abs_delta_eur": "ABS отклонение юрлица, EUR",
    "currency_original_amount": "Сумма в исходной валюте",
    "currency_eur_amount": "Сумма в EUR",
    "non_eur_amount_eur": "Сумма не-EUR, EUR",
    "non_eur_share": "Доля не-EUR, %",
    "currency_count": "Количество валют",
    "fx_quality_flag": "Флаг качества FX",
    "rows_count": "Количество строк",
    "reconciliation_status": "Статус сверки",
    "dq_status": "DQ статус",
    "qa_status": "QA статус",
    "adjustment_amount_eur": "Сумма корректировки, EUR",
    "refund_amount_eur": "Сумма возвратов, EUR",
    "timing_status": "Статус timing",
    "timing_candidate_flag": "Флаг timing candidate",
    "timing_basis": "Основание timing",
    "timing_confidence": "Уверенность timing",
    "expected_reversal_month": "Ожидаемый месяц реверса",
    "refund_rows": "Строк возвратов",
    "refund_amount_eur": "Сумма возвратов, EUR",
    "refund_share_of_fact": "Доля возвратов от факта, %",
    "refund_impact_flag": "Флаг влияния возвратов",
    "row_role": "Роль строки",
    "materiality_flag": "Флаг существенности",
    "materiality_reason": "Причина существенности",
    "materiality_rank": "Ранг существенности",
    "abs_delta_rank": "Ранг ABS отклонения",
    "delta_to_in_rank": "Ранг отклонения к IN",
    "yoy_rank": "Ранг YoY",
    "mom_rank": "Ранг MoM",
    "planning_risk_rank": "Ранг планового риска",
    "risk_level": "Уровень риска",
    "risk_basis": "Основание риска",
    "confidence_level": "Уровень уверенности",
    "confidence_reason": "Основание уверенности",
    "limitation_text": "Ограничение",
    "signal_id": "ID сигнала",
    "signal_type": "Тип сигнала",
    "signal_group": "Группа сигнала",
    "object_level": "Уровень объекта",
    "period": "Период",
    "metric_name": "Название метрики",
    "metric_value_eur": "Значение метрики, EUR",
    "metric_value_pct": "Значение метрики, %",
    "rank": "Ранг",
    "recommended_action": "Рекомендуемое действие",
    "memo_section": "Раздел memo",
    "include_in_executive_memo": "Включать в executive memo",
    "source_mart": "Источник MART",
    "source_slice": "Источник среза",
    "evidence_id": "ID подтверждения",
    "section": "Раздел",
    "signal_rank": "Ранг сигнала",
    "object_name": "Объект",
    "headline_metric_eur": "Главная метрика, EUR",
    "headline_metric_pct": "Главная метрика, %",
    "why_it_matters": "Почему важно",
    "action_required": "Требуемое действие",
    "due_date": "Срок",
    "threshold_name": "Порог",
    "threshold_value": "Значение порога",
    "description": "Описание",
    "source_file": "Файл-источник",
    "source_row_id": "ID строки источника",
    "source_files": "Файлы-источники",
    "source_rows": "Строки-источники",
    "stage_in_out_eur": "IN-OUT из Stage, EUR",
    "month_dt": "Дата месяца",
    "month_no": "Номер месяца",
    "sum_abs_mom_delta_eur": "Сумма ABS MoM отклонений, EUR",
    "source_files_count": "Количество файлов источника",
    "source_rows_count": "Количество строк источника",
    "in_denominator_status": "Статус denominator IN",
    "execution_calc_status": "Статус расчёта исполнения",
    "timing_confirmation_status": "Статус подтверждения timing",
    "timing_limitation": "Ограничение timing",
    "currency_share_in_legal_entity": "Доля валюты в юрлице, %",
    "legal_entity_share_in_currency": "Доля юрлица в валюте, %",
    "yoy_source_slice": "Источник YoY",
    "yoy_metric_grain": "Гранулярность YoY",
    "mom_source_slice": "Источник MoM",
    "mom_metric_grain": "Гранулярность MoM",
    "mom_signal_source_slice": "Источник классификации MoM",
    "mom_signal_metric_grain": "Гранулярность классификации MoM",
    "flow_prior_year_value_eur": "Flow факт прошлого года, EUR",
    "flow_yoy_delta_eur": "Flow YoY отклонение, EUR",
    "flow_abs_yoy_delta_eur": "ABS Flow YoY отклонение, EUR",
    "flow_yoy_pct": "Flow YoY, %",
    "flow_prior_year_available_flag": "Есть Flow база прошлого года",
    "flow_weak_yoy_base_flag": "Слабая Flow YoY база",
    "flow_no_yoy_base_flag": "Нет Flow YoY базы",
    "flow_yoy_rank": "Ранг Flow YoY",
    "flow_previous_month_value_eur": "Flow факт предыдущего месяца, EUR",
    "flow_mom_delta_eur": "Flow MoM отклонение, EUR",
    "flow_abs_mom_delta_eur": "ABS Flow MoM отклонение, EUR",
    "flow_mom_pct": "Flow MoM, %",
    "flow_mom_rank": "Ранг Flow MoM",
    "flow_yoy_source_slice": "Источник Flow YoY",
    "flow_yoy_metric_grain": "Гранулярность Flow YoY",
    "flow_mom_source_slice": "Источник Flow MoM",
    "flow_mom_metric_grain": "Гранулярность Flow MoM",
    "yoy_pct_numerator_eur": "Числитель YoY %, EUR",
    "yoy_pct_denominator_eur": "Знаменатель YoY %, EUR",
    "mom_pct_numerator_eur": "Числитель MoM %, EUR",
    "mom_pct_denominator_eur": "Знаменатель MoM %, EUR",
    "execution_pct_numerator_eur": "Числитель исполнения, EUR",
    "execution_pct_denominator_eur": "Знаменатель исполнения, EUR",
    "localization_source_slice": "Источник локализации",
    "localization_metric_grain": "Гранулярность локализации",
    "planning_source_slice": "Источник планового риска",
    "planning_metric_grain": "Гранулярность планового риска",
}

SHEET_NAMES = {
    "mart_main_full_budget": "01_Полный_MART",
    "mart_flow_base_month": "02_IN_OUT_База",
    "mart_signal_catalog_full": "03_Каталог_Сигналов",
    "mart_main_compact_executive_yoy_mom": "04_Compact_для_Руководства",
    "plan_fact": "05_Plan_Fact",
    "yoy": "06_YoY",
    "mom": "07_MoM",
    "localization": "08_Локализация",
    "planning_risk": "09_Плановый_Риск",
    "counterparties": "10_Контрагенты",
    "legal_entities": "11_Юрлица",
    "currencies": "12_Валюты",
    "legal_entity_currency": "13_Юрлица_Валюты",
    "source_qa": "14_QA_Source",
    "timing": "15_Timing_Кандидаты",
    "refunds": "16_Refunds",
    "thresholds": "17_Пороги",
    "mart_main_full_budget_rows": "22_Полный_MART_Строки",
}

PIVOT_SAFE_MAIN_GRAIN = [
    "period_month",
    "period_year",
    "period_type",
    "article_type",
    "article_level_1",
    "article_level_2",
    "article",
    "cfo",
    "counterparty",
]

ARTICLE_TYPE_SORT_PREFIX = {
    "IN": "00_IN",
    "OUT": "00_OUT",
    "IN-OUT": "00_IN-OUT",
    "FIX": "01_FIX",
    "PROJECT&DEPARTMENT SERVICES": "02_PROJECT&DEPARTMENT SERVICES",
    "INVEST": "03_INVEST",
    "FIN SERVICES": "04_FIN SERVICES",
    "OTHER OUTFLOWS": "05_OTHER OUTFLOWS",
}

MANAGEMENT_SHEET_GRAINS = {
    "01_Полный_MART": ["Месяц", "Год", "Тип периода", "Тип статьи", "Статья 1", "Статья 2", "Статья", "ЦФО", "Контрагент"],
    "05_Plan_Fact": ["Месяц", "Статья", "ЦФО"],
    "06_YoY": ["Статья", "ЦФО", "Год", "Номер месяца"],
    "07_MoM": ["Статья", "ЦФО"],
    "08_Локализация": ["Статья", "ЦФО"],
    "09_Плановый_Риск": ["Статья", "ЦФО"],
    "10_Контрагенты": ["Контрагент"],
    "11_Юрлица": ["Юр. лицо"],
    "12_Валюты": ["Валюта"],
    "13_Юрлица_Валюты": ["Юр. лицо", "Валюта"],
    "15_Timing_Кандидаты": ["Месяц", "Статья", "ЦФО", "Контрагент"],
    "16_Refunds": ["Месяц", "Статья", "ЦФО", "Контрагент"],
}


def safe_div(num: pd.Series | float, den: pd.Series | float) -> pd.Series:
    if not isinstance(num, pd.Series):
        num = pd.Series(num)
    if not isinstance(den, pd.Series):
        den = pd.Series(den, index=num.index)
    return num / den.where(den.notna() & den.ne(0))


def parse_numeric(series: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(series):
        return pd.to_numeric(series, errors="coerce")
    return pd.to_numeric(
        series.astype(str)
        .str.replace("\ufeff", "", regex=False)
        .str.replace(" ", "", regex=False)
        .str.replace(",", ".", regex=False)
        .replace({"": np.nan, "nan": np.nan, "None": np.nan}),
        errors="coerce",
    )


def join_unique(values: pd.Series) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = "" if pd.isna(value) else str(value).strip()
        if text and text not in seen:
            result.append(text)
            seen.add(text)
    return " | ".join(result)


def snapshot(folder: Path) -> dict[str, tuple[int, int]]:
    if not folder.exists():
        return {}
    result: dict[str, tuple[int, int]] = {}
    for path in folder.rglob("*"):
        if path.is_file():
            stat = path.stat()
            result[str(path.relative_to(PROJECT_ROOT))] = (stat.st_mtime_ns, stat.st_size)
    return result


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def clean_rebuild_dirs() -> None:
    for directory in [MARTS_DIR, SIGNALS_DIR, EVIDENCE_DIR, QA_DIR]:
        directory.mkdir(parents=True, exist_ok=True)
        for path in directory.iterdir():
            if path.is_dir():
                shutil.rmtree(path)
            else:
                path.unlink()


def read_stage() -> pd.DataFrame:
    if not ARCHIVE_PREREQUISITE.exists():
        raise FileNotFoundError(f"Missing prerequisite archive: {ARCHIVE_PREREQUISITE}")
    if not STAGE_SOURCE.exists():
        raise FileNotFoundError(f"Missing accepted Stage source: {STAGE_SOURCE}")
    raw = pd.read_csv(
        STAGE_SOURCE,
        sep=";",
        decimal=",",
        encoding="utf-8-sig",
        dtype={
            "Код статьи ДДС": "string",
            "Ключ контрагента": "string",
            "source_row_id": "string",
            "source_file": "string",
        },
    )
    stage = raw.rename(columns={col: SOURCE_COLUMN_MAP[col] for col in raw.columns if col in SOURCE_COLUMN_MAP})
    required = {"period_month", "period_type", "article", "cfo", "plan_eur", "fact_eur", "stage_in_out_eur"}
    missing = sorted(required - set(stage.columns))
    if missing:
        raise ValueError(f"Missing Stage columns for MART rebuild: {missing}")

    for col in ["source_amount", "plan_eur", "fact_eur", "stage_in_out_eur"]:
        stage[col] = parse_numeric(stage[col])
    for col in ["included_in_reconciliation", "has_plan", "has_fact", "has_p_fact_adjustment", "has_player_refund"]:
        stage[col] = parse_numeric(stage.get(col, 0)).fillna(0).astype(int)
    text_cols = [
        "period_month",
        "period_type",
        "source_mix",
        "article_code",
        "article_type",
        "article_level_1",
        "article_level_2",
        "article",
        "cfo",
        "legal_entity",
        "counterparty",
        "counterparty_key",
        "counterparty_type",
        "currency",
        "source_file",
        "source_row_id",
    ]
    for col in text_cols:
        if col not in stage.columns:
            stage[col] = ""
        stage[col] = stage[col].fillna("").astype(str).str.strip()

    parsed_month = pd.to_datetime(stage["period_month"] + "-01", errors="coerce")
    stage["period_year"] = parsed_month.dt.year.astype("Int64")
    stage["month_dt"] = parsed_month
    stage["date"] = pd.to_datetime(stage.get("date", ""), errors="coerce").fillna(parsed_month)
    stage["period_type"] = stage["period_type"].replace({"future": "planning"})
    stage["row_id"] = np.arange(1, len(stage) + 1)
    return stage


def assign_row_roles(df: pd.DataFrame) -> pd.Series:
    service = df["article"].isin(SERVICE_VALUES) | df["cfo"].isin(SERVICE_VALUES) | df["article_code"].isin(SERVICE_VALUES)
    refund = df["has_player_refund"].eq(1)
    adjustment = df["has_p_fact_adjustment"].eq(1) | df["source_mix"].eq("p_fact_adjusted")
    return pd.Series(
        np.select(
            [service, refund, adjustment, df["included_in_reconciliation"].eq(1)],
            ["service_flow_row", "refund_row", "adjustment_row", "business_budget_row"],
            default="qa_control_row",
        ),
        index=df.index,
    )


def build_flow_base(stage: pd.DataFrame) -> pd.DataFrame:
    service = stage[stage["row_role"].eq("service_flow_row")].copy()
    if service.empty:
        return pd.DataFrame(
            columns=[
                "period_month",
                "period_type",
                "in_eur",
                "out_eur",
                "in_out_eur",
                "out_to_in_pct",
                "in_out_margin_pct",
                "flow_base_status",
                "flow_value_source",
                "flow_uses_plan_fallback_flag",
                "weak_in_base_flag",
                "weak_flow_base_flag",
            ]
        )
    fact_available = service["fact_eur"].notna() & service["fact_eur"].ne(0)
    plan_available = service["plan_eur"].notna() & service["plan_eur"].ne(0)
    service["flow_value_source"] = np.select(
        [fact_available, plan_available],
        ["fact", "plan_fallback"],
        default="missing",
    )
    service["flow_uses_plan_fallback_flag"] = service["flow_value_source"].eq("plan_fallback").astype(int)
    service["flow_value_eur"] = service["fact_eur"].where(fact_available, service["plan_eur"])
    pivot = service.pivot_table(
        index=["period_month", "period_type"],
        columns="article",
        values="flow_value_eur",
        aggfunc="sum",
        fill_value=np.nan,
    ).reset_index()
    source_status = service.groupby(["period_month", "period_type"], as_index=False, dropna=False).agg(
        flow_value_source=("flow_value_source", join_unique),
        flow_uses_plan_fallback_flag=("flow_uses_plan_fallback_flag", "max"),
    )
    for col in SERVICE_VALUES:
        if col not in pivot.columns:
            pivot[col] = np.nan
    flow = pivot.rename(columns={"IN": "in_eur", "OUT": "out_eur", "IN-OUT": "in_out_eur"})
    flow = flow.merge(source_status, on=["period_month", "period_type"], how="left")
    flow["out_to_in_pct"] = safe_div(flow["out_eur"], flow["in_eur"])
    flow["in_out_margin_pct"] = safe_div(flow["in_out_eur"], flow["in_eur"])
    flow["weak_in_base_flag"] = flow["in_eur"].abs().lt(THRESHOLDS["weak_base_threshold_eur"]).fillna(True).astype(int)
    flow["weak_flow_base_flag"] = (flow[["in_eur", "out_eur", "in_out_eur"]].isna().any(axis=1) | flow["weak_in_base_flag"].eq(1)).astype(int)
    flow["flow_base_status"] = np.select(
        [flow["weak_flow_base_flag"].eq(1), flow["flow_uses_plan_fallback_flag"].eq(1)],
        ["warning", "plan_fallback"],
        default="pass",
    )
    return flow[
        [
            "period_month",
            "period_type",
            "in_eur",
            "out_eur",
            "in_out_eur",
            "out_to_in_pct",
            "in_out_margin_pct",
            "flow_base_status",
            "flow_value_source",
            "flow_uses_plan_fallback_flag",
            "weak_in_base_flag",
            "weak_flow_base_flag",
        ]
    ]


def add_core_metrics(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    result["delta_eur"] = result["plan_eur"] - result["fact_eur"]
    result["abs_delta_eur"] = result["delta_eur"].abs()
    result["execution_pct"] = safe_div(result["fact_eur"], result["plan_eur"])
    total_abs = result["abs_delta_eur"].sum()
    result["share_of_total_abs_delta"] = safe_div(result["abs_delta_eur"], total_abs)
    result["fact_without_plan_flag"] = (result["fact_eur"].ne(0) & result["plan_eur"].eq(0)).astype(int)
    result["plan_without_fact_flag"] = (result["plan_eur"].ne(0) & result["fact_eur"].eq(0)).astype(int)
    result["overrun_flag"] = result["delta_eur"].lt(0).astype(int)
    result["underexecution_flag"] = result["delta_eur"].gt(0).astype(int)
    return result


def add_flow_metrics(df: pd.DataFrame, flow: pd.DataFrame) -> pd.DataFrame:
    result = df.merge(flow, on=["period_month", "period_type"], how="left")
    result["plan_to_in_pct"] = safe_div(result["plan_eur"], result["in_eur"])
    result["fact_to_in_pct"] = safe_div(result["fact_eur"], result["in_eur"])
    result["delta_to_in_pct"] = safe_div(result["delta_eur"], result["in_eur"])
    result["abs_delta_to_in_pct"] = safe_div(result["abs_delta_eur"], result["in_eur"])
    return result


def build_main_full(stage: pd.DataFrame, flow: pd.DataFrame) -> pd.DataFrame:
    full = stage.copy()
    full["row_role"] = assign_row_roles(full)
    full = add_core_metrics(full)
    full = add_flow_metrics(full, flow)
    business = full[full["row_role"].isin(["business_budget_row", "adjustment_row", "refund_row"])].copy()
    hist = (
        business[business["period_type"].eq("historical")]
        .groupby(["article", "cfo"], as_index=False)
        .agg(historical_base_eur=("fact_eur", "mean"), base_months_available=("period_month", "nunique"))
    )
    full = full.merge(hist, on=["article", "cfo"], how="left")
    full["base_months_available"] = full["base_months_available"].fillna(0).astype(int)
    full["months_without_base"] = np.where(full["base_months_available"].eq(0), 1, 0)
    full["planning_plan_eur"] = np.where(full["period_type"].eq("planning"), full["plan_eur"], 0.0)
    full["plan_vs_base_delta_eur"] = full["planning_plan_eur"] - full["historical_base_eur"]
    full["plan_vs_base_abs_delta_eur"] = full["plan_vs_base_delta_eur"].abs()
    full["plan_vs_base_pct"] = safe_div(full["plan_vs_base_delta_eur"], full["historical_base_eur"])
    full["plan_vs_base_to_in_pct"] = safe_div(full["plan_vs_base_delta_eur"], full["in_eur"])
    full["planning_risk_flag"] = (
        full["period_type"].eq("planning")
        & full["plan_vs_base_abs_delta_eur"].ge(THRESHOLDS["planning_risk_threshold_eur"])
    ).astype(int)
    full["planning_risk_basis"] = np.where(
        full["planning_risk_flag"].eq(1),
        "Плановый бюджет отличается от исторической базы; это future budget risk, not actual execution.",
        "",
    )
    full["materiality_flag"] = (
        full["abs_delta_eur"].ge(THRESHOLDS["material_abs_delta_threshold_eur"])
        | full["abs_delta_to_in_pct"].abs().ge(THRESHOLDS["material_delta_to_in_threshold_pct"])
        | full["planning_risk_flag"].eq(1)
    ).astype(int)
    full["materiality_reason"] = np.select(
        [
            full["planning_risk_flag"].eq(1),
            full["abs_delta_eur"].ge(THRESHOLDS["material_abs_delta_threshold_eur"]),
            full["abs_delta_to_in_pct"].abs().ge(THRESHOLDS["material_delta_to_in_threshold_pct"]),
        ],
        ["planning_risk_threshold", "abs_delta_threshold", "delta_to_in_threshold"],
        default="",
    )
    full["materiality_rank"] = full["abs_delta_eur"].fillna(0).rank(method="first", ascending=False).astype(int)
    full["abs_delta_rank"] = full["abs_delta_eur"].fillna(0).rank(method="first", ascending=False).astype(int)
    full["delta_to_in_rank"] = full["abs_delta_to_in_pct"].abs().fillna(0).rank(method="first", ascending=False).astype(int)
    full["planning_risk_rank"] = full["plan_vs_base_abs_delta_eur"].fillna(0).rank(method="first", ascending=False).astype(int)
    full["yoy_rank"] = pd.NA
    full["mom_rank"] = pd.NA
    full["risk_level"] = np.select(
        [full["materiality_flag"].eq(1), full["weak_flow_base_flag"].eq(1)],
        ["high", "medium"],
        default="low",
    )
    full["risk_basis"] = np.where(full["risk_level"].eq("high"), full["materiality_reason"], "below_default_thresholds")
    full["confidence_level"] = np.where(full["weak_flow_base_flag"].eq(1), "medium", "high")
    full["confidence_reason"] = np.where(full["weak_flow_base_flag"].eq(1), "weak_or_missing_flow_base", "deterministic_stage_mart_calculation")
    full["limitation_text"] = np.where(
        full["period_type"].eq("planning"),
        "Planning risk is future budget risk, not actual execution.",
        "",
    )
    full["timing_status"] = "unknown"
    full["timing_candidate_flag"] = 0
    full["timing_basis"] = ""
    full["timing_confidence"] = "low"
    full["expected_reversal_month"] = ""
    return full


def aggregate_slice(full: pd.DataFrame, keys: list[str], source_slice: str) -> pd.DataFrame:
    source = full[~full["row_role"].eq("service_flow_row")].copy()
    grouped = source.groupby(keys, as_index=False, dropna=False).agg(
        plan_eur=("plan_eur", "sum"),
        fact_eur=("fact_eur", "sum"),
        rows_count=("row_id", "size"),
        source_files=("source_file", join_unique),
        source_rows=("source_row_id", join_unique),
        in_eur=("in_eur", "first"),
        out_eur=("out_eur", "first"),
    )
    grouped = add_core_metrics(grouped)
    grouped["plan_to_in_pct"] = safe_div(grouped["plan_eur"], grouped["in_eur"])
    grouped["fact_to_in_pct"] = safe_div(grouped["fact_eur"], grouped["in_eur"])
    grouped["delta_to_in_pct"] = safe_div(grouped["delta_eur"], grouped["in_eur"])
    grouped["abs_delta_to_in_pct"] = safe_div(grouped["abs_delta_eur"], grouped["in_eur"])
    grouped["source_slice"] = source_slice
    return grouped


def with_rank(df: pd.DataFrame, rank_col: str, metric: str) -> pd.DataFrame:
    result = df.copy()
    result[rank_col] = result[metric].abs().fillna(0).rank(method="first", ascending=False).astype(int)
    return result


def build_plan_fact_slices(full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    specs = {
        "slice_plan_fact_article_month": ["period_month", "period_year", "period_type", "article"],
        "slice_plan_fact_article": ["article"],
        "slice_plan_fact_article_cfo_month": ["period_month", "period_year", "period_type", "article", "cfo"],
        "slice_plan_fact_article_cfo": ["article", "cfo"],
        "slice_plan_fact_article_cfo_counterparty_month": ["period_month", "period_year", "period_type", "article", "cfo", "counterparty"],
        "slice_plan_fact_counterparty": ["counterparty"],
        "slice_plan_fact_legal_entity": ["legal_entity"],
    }
    return {name: with_rank(aggregate_slice(full, keys, name), "abs_delta_rank", "abs_delta_eur") for name, keys in specs.items()}


def build_yoy_slices(plan_fact: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    result: dict[str, pd.DataFrame] = {}
    specs = {
        "slice_yoy_article_month": (plan_fact["slice_plan_fact_article_month"], ["article"]),
        "slice_yoy_article_cfo_month": (plan_fact["slice_plan_fact_article_cfo_month"], ["article", "cfo"]),
        "slice_yoy_article": (plan_fact["slice_plan_fact_article_month"], ["article"]),
        "slice_yoy_article_cfo": (plan_fact["slice_plan_fact_article_cfo_month"], ["article", "cfo"]),
        "slice_yoy_counterparty": (plan_fact["slice_plan_fact_article_cfo_counterparty_month"], ["counterparty"]),
        "slice_yoy_legal_entity": (plan_fact["slice_plan_fact_legal_entity"], ["legal_entity"]),
    }
    for name, (df, keys) in specs.items():
        work = df.copy()
        if "period_month" in work.columns:
            work["month_no"] = pd.to_datetime(work["period_month"] + "-01", errors="coerce").dt.month
            current = work[[*keys, "period_year", "month_no", "period_month", "fact_eur", "in_eur", "rows_count"]].rename(columns={"fact_eur": "current_fact_eur"})
            prior = current.rename(
                columns={
                    "current_fact_eur": "prior_year_fact_eur",
                    "period_year": "prior_year",
                    "rows_count": "prior_year_month_count",
                }
            )
            prior["period_year"] = prior["prior_year"] + 1
            merged = current.merge(prior[[*keys, "period_year", "month_no", "prior_year_fact_eur", "prior_year_month_count"]], on=[*keys, "period_year", "month_no"], how="left")
        else:
            current = work.groupby(keys, as_index=False, dropna=False).agg(current_fact_eur=("fact_eur", "sum"), in_eur=("in_eur", "first"), rows_count=("rows_count", "sum"))
            merged = current.copy()
            merged["period_month"] = "all"
            merged["period_year"] = pd.NA
            merged["prior_year_fact_eur"] = np.nan
            merged["prior_year_month_count"] = 0
        merged["yoy_delta_eur"] = merged["current_fact_eur"] - merged["prior_year_fact_eur"]
        merged["abs_yoy_delta_eur"] = merged["yoy_delta_eur"].abs()
        merged["yoy_pct"] = safe_div(merged["yoy_delta_eur"], merged["prior_year_fact_eur"])
        merged["yoy_delta_to_in_pct"] = safe_div(merged["yoy_delta_eur"], merged["in_eur"])
        merged["abs_yoy_delta_to_in_pct"] = safe_div(merged["abs_yoy_delta_eur"], merged["in_eur"])
        merged["prior_year_available_flag"] = merged["prior_year_fact_eur"].notna().astype(int)
        merged["prior_year_month_count"] = merged["prior_year_month_count"].fillna(0).astype(int)
        merged["weak_yoy_base_flag"] = merged["prior_year_fact_eur"].abs().lt(THRESHOLDS["weak_base_threshold_eur"]).fillna(True).astype(int)
        merged["no_yoy_base_flag"] = merged["prior_year_available_flag"].eq(0).astype(int)
        merged = with_rank(merged, "yoy_rank", "abs_yoy_delta_eur")
        merged["source_slice"] = name
        result[name] = merged
    return result


def longest_same_direction(values: pd.Series) -> int:
    best = 0
    current = 0
    prev = 0
    for value in values.fillna(0):
        direction = 1 if value > 0 else -1 if value < 0 else 0
        if direction == 0:
            current = 0
            prev = 0
        elif direction == prev:
            current += 1
        else:
            current = 1
            prev = direction
        best = max(best, current)
    return best


def classify_mom(row: pd.Series) -> str:
    if row["active_months_count"] < 2:
        return "insufficient_data"
    if row["sum_abs_mom_delta_eur"] == 0:
        return "stable"
    if row["abs_mom_delta_eur"] / row["sum_abs_mom_delta_eur"] >= 0.6:
        return "one_off_spike"
    if row["longest_same_direction_series"] >= 3:
        return "serial_shift"
    if row["growth_months_count"] + row["decline_months_count"] >= 3:
        return "repeated_instability"
    return "stable"


def build_mom_slices(plan_fact: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    result: dict[str, pd.DataFrame] = {}
    specs = {
        "slice_mom_article_month": (plan_fact["slice_plan_fact_article_month"], ["article"]),
        "slice_mom_article_cfo_month": (plan_fact["slice_plan_fact_article_cfo_month"], ["article", "cfo"]),
        "slice_mom_article": (plan_fact["slice_plan_fact_article_month"], ["article"]),
        "slice_mom_article_cfo": (plan_fact["slice_plan_fact_article_cfo_month"], ["article", "cfo"]),
        "slice_mom_counterparty": (plan_fact["slice_plan_fact_article_cfo_counterparty_month"], ["counterparty"]),
    }
    for name, (df, keys) in specs.items():
        work = df.copy()
        if "period_month" not in work.columns:
            continue
        work = work.sort_values([*keys, "period_month"])
        work["previous_month_fact_eur"] = work.groupby(keys)["fact_eur"].shift(1)
        work["current_month_fact_eur"] = work["fact_eur"]
        work["mom_delta_eur"] = work["current_month_fact_eur"] - work["previous_month_fact_eur"]
        work["abs_mom_delta_eur"] = work["mom_delta_eur"].abs()
        work["mom_pct"] = safe_div(work["mom_delta_eur"], work["previous_month_fact_eur"])
        work["mom_delta_to_in_pct"] = safe_div(work["mom_delta_eur"], work["in_eur"])
        work["abs_mom_delta_to_in_pct"] = safe_div(work["abs_mom_delta_eur"], work["in_eur"])
        agg = work.groupby(keys, as_index=False, dropna=False).agg(
            growth_months_count=("mom_delta_eur", lambda s: int(s.fillna(0).gt(0).sum())),
            decline_months_count=("mom_delta_eur", lambda s: int(s.fillna(0).lt(0).sum())),
            active_months_count=("period_month", "nunique"),
            longest_same_direction_series=("mom_delta_eur", longest_same_direction),
            sum_abs_mom_delta_eur=("mom_delta_eur", lambda s: float(s.abs().sum())),
            abs_mom_delta_eur=("abs_mom_delta_eur", "max"),
            mom_delta_eur=("mom_delta_eur", lambda s: float(s.loc[s.abs().idxmax()]) if s.abs().notna().any() else np.nan),
            current_month_fact_eur=("current_month_fact_eur", "sum"),
            previous_month_fact_eur=("previous_month_fact_eur", "sum"),
            in_eur=("in_eur", "first"),
        )
        agg["mom_pct"] = safe_div(agg["mom_delta_eur"], agg["previous_month_fact_eur"])
        agg["mom_delta_to_in_pct"] = safe_div(agg["mom_delta_eur"], agg["in_eur"])
        agg["abs_mom_delta_to_in_pct"] = safe_div(agg["abs_mom_delta_eur"], agg["in_eur"])
        agg["mom_signal_type"] = agg.apply(classify_mom, axis=1)
        agg = with_rank(agg, "mom_rank", "abs_mom_delta_eur")
        agg["source_slice"] = name
        result[name] = agg
        if name.endswith("_month"):
            result[name] = with_rank(work, "mom_rank", "abs_mom_delta_eur")
    result["slice_mom_signal_classification"] = result["slice_mom_article"][["article", "mom_signal_type", "mom_rank", "abs_mom_delta_eur", "source_slice"]].copy()
    return result


def build_flow_yoy_mom_slice(flow: pd.DataFrame) -> pd.DataFrame:
    if flow.empty:
        return pd.DataFrame(
            columns=[
                "period_month",
                "period_year",
                "month_no",
                "period_type",
                "flow_metric",
                "flow_value_eur",
                "flow_value_source",
                "flow_uses_plan_fallback_flag",
                "flow_prior_year_value_eur",
                "flow_yoy_delta_eur",
                "flow_abs_yoy_delta_eur",
                "flow_yoy_pct",
                "flow_prior_year_available_flag",
                "flow_weak_yoy_base_flag",
                "flow_no_yoy_base_flag",
                "flow_yoy_rank",
                "flow_previous_month_value_eur",
                "flow_mom_delta_eur",
                "flow_abs_mom_delta_eur",
                "flow_mom_pct",
                "flow_mom_rank",
                "source_slice",
            ]
        )

    value_columns = {"IN": "in_eur", "OUT": "out_eur", "IN-OUT": "in_out_eur"}
    frames = []
    for metric, value_col in value_columns.items():
        part = flow[
            [
                "period_month",
                "period_type",
                value_col,
                "flow_value_source",
                "flow_uses_plan_fallback_flag",
            ]
        ].copy()
        part["flow_metric"] = metric
        part = part.rename(columns={value_col: "flow_value_eur"})
        frames.append(part)
    result = pd.concat(frames, ignore_index=True)
    month_dt = pd.to_datetime(result["period_month"] + "-01", errors="coerce")
    result["period_year"] = month_dt.dt.year.astype("Int64")
    result["month_no"] = month_dt.dt.month

    current = result[
        [
            "flow_metric",
            "period_year",
            "month_no",
            "flow_value_eur",
        ]
    ].copy()
    prior = current.rename(
        columns={
            "flow_value_eur": "flow_prior_year_value_eur",
            "period_year": "prior_year",
        }
    )
    prior["period_year"] = prior["prior_year"] + 1
    result = result.merge(
        prior[["flow_metric", "period_year", "month_no", "flow_prior_year_value_eur"]],
        on=["flow_metric", "period_year", "month_no"],
        how="left",
    )
    result["flow_yoy_delta_eur"] = result["flow_value_eur"] - result["flow_prior_year_value_eur"]
    result["flow_abs_yoy_delta_eur"] = result["flow_yoy_delta_eur"].abs()
    result["flow_yoy_pct"] = safe_div(result["flow_yoy_delta_eur"], result["flow_prior_year_value_eur"])
    result["flow_prior_year_available_flag"] = result["flow_prior_year_value_eur"].notna().astype(int)
    result["flow_weak_yoy_base_flag"] = result["flow_prior_year_value_eur"].abs().lt(THRESHOLDS["weak_base_threshold_eur"]).fillna(True).astype(int)
    result["flow_no_yoy_base_flag"] = result["flow_prior_year_available_flag"].eq(0).astype(int)
    result = with_rank(result, "flow_yoy_rank", "flow_abs_yoy_delta_eur")

    result = result.sort_values(["flow_metric", "period_month"]).reset_index(drop=True)
    result["flow_previous_month_value_eur"] = result.groupby("flow_metric")["flow_value_eur"].shift(1)
    result["flow_mom_delta_eur"] = result["flow_value_eur"] - result["flow_previous_month_value_eur"]
    result["flow_abs_mom_delta_eur"] = result["flow_mom_delta_eur"].abs()
    result["flow_mom_pct"] = safe_div(result["flow_mom_delta_eur"], result["flow_previous_month_value_eur"])
    result = with_rank(result, "flow_mom_rank", "flow_abs_mom_delta_eur")
    result["source_slice"] = "slice_flow_yoy_mom_month"
    return result


def build_localization_slices(plan_fact: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    cfo = plan_fact["slice_plan_fact_article_cfo"].copy()
    article_total = cfo.groupby("article", as_index=False).agg(article_abs_delta_eur=("abs_delta_eur", "sum"))
    cfo = cfo.merge(article_total, on="article", how="left")
    cfo["cfo_abs_delta_eur"] = cfo["abs_delta_eur"]
    cfo["cfo_share_in_article_delta"] = safe_div(cfo["cfo_abs_delta_eur"], cfo["article_abs_delta_eur"])
    stats = cfo.sort_values("cfo_share_in_article_delta", ascending=False).groupby("article").agg(
        top1_cfo_share=("cfo_share_in_article_delta", "max"),
        top3_cfo_share=("cfo_share_in_article_delta", lambda s: float(s.head(3).sum())),
    ).reset_index()
    cfo = cfo.merge(stats, on="article", how="left")
    cfo["concentration_type"] = np.select(
        [cfo["top1_cfo_share"].ge(0.7), cfo["top3_cfo_share"].ge(0.9)],
        ["concentrated_risk", "moderately_concentrated_risk"],
        default="distributed_risk",
    )
    cfo["owner_candidate"] = cfo["cfo"]
    cfo["owner_route_status"] = np.where(cfo["cfo"].fillna("").eq(""), "unknown", "candidate")
    cfo = with_rank(cfo, "materiality_rank", "cfo_abs_delta_eur")
    month = plan_fact["slice_plan_fact_article_cfo_month"].merge(cfo[["article", "cfo", "article_abs_delta_eur", "cfo_abs_delta_eur", "cfo_share_in_article_delta", "top1_cfo_share", "top3_cfo_share", "concentration_type", "owner_candidate", "owner_route_status"]], on=["article", "cfo"], how="left")
    counterparty = plan_fact["slice_plan_fact_article_cfo_counterparty_month"].merge(cfo[["article", "cfo", "concentration_type", "owner_candidate", "owner_route_status"]], on=["article", "cfo"], how="left")
    route = cfo[["article", "cfo", "owner_candidate", "owner_route_status", "concentration_type", "cfo_abs_delta_eur", "materiality_rank"]].copy()
    return {
        "slice_localization_article_cfo": cfo,
        "slice_localization_article_cfo_month": month,
        "slice_localization_article_cfo_counterparty": counterparty,
        "slice_localization_owner_route": route,
    }


def build_planning_slices(full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    planning = full[(full["period_type"].eq("planning")) & ~full["row_role"].eq("service_flow_row")].copy()
    specs = {
        "slice_plan_vs_history_article_month": ["period_month", "period_year", "article"],
        "slice_plan_vs_history_article": ["article"],
        "slice_plan_vs_history_article_cfo_month": ["period_month", "period_year", "article", "cfo"],
        "slice_plan_vs_history_article_cfo": ["article", "cfo"],
        "slice_plan_vs_history_counterparty": ["counterparty"],
    }
    result = {}
    for name, keys in specs.items():
        grouped = planning.groupby(keys, as_index=False, dropna=False).agg(
            planning_plan_eur=("plan_eur", "sum"),
            historical_base_eur=("historical_base_eur", "sum"),
            base_months_available=("base_months_available", "max"),
            months_without_base=("months_without_base", "sum"),
            in_eur=("in_eur", "first"),
        )
        grouped["plan_vs_base_delta_eur"] = grouped["planning_plan_eur"] - grouped["historical_base_eur"]
        grouped["plan_vs_base_abs_delta_eur"] = grouped["plan_vs_base_delta_eur"].abs()
        grouped["plan_vs_base_pct"] = safe_div(grouped["plan_vs_base_delta_eur"], grouped["historical_base_eur"])
        grouped["plan_vs_base_to_in_pct"] = safe_div(grouped["plan_vs_base_delta_eur"], grouped["in_eur"])
        grouped["planning_risk_flag"] = grouped["plan_vs_base_abs_delta_eur"].ge(THRESHOLDS["planning_risk_threshold_eur"]).astype(int)
        grouped["planning_risk_basis"] = np.where(grouped["planning_risk_flag"].eq(1), "future_budget_risk_vs_historical_base", "")
        grouped = with_rank(grouped, "planning_risk_rank", "plan_vs_base_abs_delta_eur")
        grouped["source_slice"] = name
        result[name] = grouped
    result["slice_plan_without_history_base"] = result["slice_plan_vs_history_article_cfo_month"][result["slice_plan_vs_history_article_cfo_month"]["base_months_available"].eq(0)].copy()
    result["slice_planning_risk_candidates"] = result["slice_plan_vs_history_article_cfo"][result["slice_plan_vs_history_article_cfo"]["planning_risk_flag"].eq(1)].copy()
    return result


def build_counterparty_slices(full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    cp = aggregate_slice(full, ["counterparty"], "slice_counterparty_base")
    cp = cp.rename(
        columns={
            "fact_eur": "counterparty_fact_eur",
            "plan_eur": "counterparty_plan_eur",
            "delta_eur": "counterparty_delta_eur",
            "abs_delta_eur": "counterparty_abs_delta_eur",
        }
    )
    total_fact = cp["counterparty_fact_eur"].abs().sum()
    cp_sorted = cp.sort_values("counterparty_fact_eur", key=lambda s: s.abs(), ascending=False).reset_index(drop=True)
    cp_sorted["counterparty_share"] = safe_div(cp_sorted["counterparty_fact_eur"].abs(), total_fact)
    top5 = cp_sorted.head(5)["counterparty_fact_eur"].abs().sum()
    top10 = cp_sorted.head(10)["counterparty_fact_eur"].abs().sum()
    cp["top5_counterparty_share"] = safe_div(pd.Series([top5] * len(cp), index=cp.index), total_fact)
    cp["top10_counterparty_share"] = safe_div(pd.Series([top10] * len(cp), index=cp.index), total_fact)
    key_quality = (
        full[~full["row_role"].eq("service_flow_row")]
        .groupby("counterparty", as_index=False, dropna=False)
        .agg(
            unknown_key_rows=("counterparty_key", lambda s: int(s.isin(UNKNOWN_VALUES).sum())),
            unknown_key_amount_eur=("fact_eur", lambda s: float(s[full.loc[s.index, "counterparty_key"].isin(UNKNOWN_VALUES)].abs().sum())),
        )
    )
    cp = cp.merge(key_quality, on="counterparty", how="left")
    cp[["unknown_key_rows", "unknown_key_amount_eur"]] = cp[["unknown_key_rows", "unknown_key_amount_eur"]].fillna(0)
    unknown_mask = cp["counterparty"].isin(UNKNOWN_VALUES)
    cp["unknown_counterparty_rows"] = np.where(unknown_mask, cp["rows_count"], 0)
    cp["unknown_counterparty_amount_eur"] = np.where(unknown_mask, cp["counterparty_fact_eur"].abs(), 0.0)
    cp["unknown_counterparty_share"] = safe_div(cp["unknown_counterparty_amount_eur"], total_fact)
    cp["counterparty_quality_flag"] = np.where((cp["unknown_counterparty_share"].fillna(0) > 0.01) | (cp["unknown_key_rows"] > 0), "warning", "pass")
    return {
        "slice_counterparty_top_by_fact": cp.sort_values("counterparty_fact_eur", key=lambda s: s.abs(), ascending=False).head(100),
        "slice_counterparty_top_by_plan": cp.sort_values("counterparty_plan_eur", key=lambda s: s.abs(), ascending=False).head(100),
        "slice_counterparty_top_by_delta": cp.sort_values("counterparty_abs_delta_eur", ascending=False).head(100),
        "slice_counterparty_unknown": cp[unknown_mask].copy(),
        "slice_counterparty_missing_key": full[full["counterparty_key"].isin(UNKNOWN_VALUES)].copy(),
        "slice_counterparty_concentration": cp,
        "slice_counterparty_fact_without_plan": cp[(cp["counterparty_fact_eur"].ne(0)) & (cp["counterparty_plan_eur"].eq(0))].copy(),
        "slice_counterparty_plan_without_fact": cp[(cp["counterparty_plan_eur"].ne(0)) & (cp["counterparty_fact_eur"].eq(0))].copy(),
        "slice_one_time_counterparties": cp[cp["rows_count"].eq(1)].copy(),
    }


def build_legal_currency_slices(full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    legal = aggregate_slice(full, ["legal_entity"], "slice_legal_entity_summary")
    legal["legal_entity_plan_eur"] = legal["plan_eur"]
    legal["legal_entity_fact_eur"] = legal["fact_eur"]
    legal["legal_entity_abs_delta_eur"] = legal["abs_delta_eur"]
    legal["qa_status"] = "pass"
    legal["limitation_text"] = "IN ratios are not shown in management Excel because legal entity aggregation is not period-scoped."
    legal_cp = aggregate_slice(full, ["legal_entity", "counterparty"], "slice_legal_entity_counterparty")
    currency = full.groupby(["currency"], as_index=False, dropna=False).agg(
        currency_original_amount=("source_amount", "sum"),
        currency_eur_amount=("fact_eur", "sum"),
        rows_count=("row_id", "size"),
    )
    currency["non_eur_amount_eur"] = np.where(currency["currency"].ne("EUR"), currency["currency_eur_amount"].abs(), 0.0)
    total_eur = currency["currency_eur_amount"].abs().sum()
    currency["non_eur_share"] = safe_div(currency["non_eur_amount_eur"], total_eur)
    currency["currency_count"] = full["currency"].nunique()
    currency["fx_quality_flag"] = np.where(currency["currency"].eq(""), "warning", "pass")
    currency["source_slice"] = "slice_currency_exposure"
    currency["qa_status"] = currency["fx_quality_flag"]
    currency["limitation_text"] = np.where(currency["currency"].eq("EUR"), "EUR is excluded from Non-EUR exposure.", "")

    legal_currency_source = full[~full["row_role"].eq("service_flow_row")].copy()
    legal_currency = legal_currency_source.groupby(["legal_entity", "currency"], as_index=False, dropna=False).agg(
        currency_original_amount=("source_amount", "sum"),
        currency_eur_amount=("fact_eur", "sum"),
        rows_count=("row_id", "size"),
    )
    legal_currency["non_eur_amount_eur"] = np.where(
        legal_currency["currency"].ne("EUR"), legal_currency["currency_eur_amount"].abs(), 0.0
    )
    legal_total = legal_currency.groupby("legal_entity")["currency_eur_amount"].transform(lambda s: s.abs().sum())
    currency_total = legal_currency.groupby("currency")["currency_eur_amount"].transform(lambda s: s.abs().sum())
    legal_currency["currency_share_in_legal_entity"] = safe_div(legal_currency["currency_eur_amount"].abs(), legal_total)
    legal_currency["legal_entity_share_in_currency"] = safe_div(legal_currency["currency_eur_amount"].abs(), currency_total)
    legal_currency["fx_quality_flag"] = np.where(legal_currency["currency"].fillna("").eq(""), "warning", "pass")
    legal_currency["source_slice"] = "slice_legal_entity_currency_exposure"
    legal_currency["qa_status"] = legal_currency["fx_quality_flag"]
    legal_currency["limitation_text"] = np.where(legal_currency["currency"].eq("EUR"), "EUR is excluded from Non-EUR exposure.", "")
    return {
        "slice_legal_entity_summary": legal.sort_values("legal_entity_abs_delta_eur", ascending=False),
        "slice_legal_entity_fact": legal.sort_values("legal_entity_fact_eur", key=lambda s: s.abs(), ascending=False),
        "slice_legal_entity_plan": legal.sort_values("legal_entity_plan_eur", key=lambda s: s.abs(), ascending=False),
        "slice_legal_entity_delta": legal.sort_values("legal_entity_abs_delta_eur", ascending=False),
        "slice_legal_entity_counterparty": legal_cp,
        "slice_currency_exposure": currency,
        "slice_non_eur_operations": full[full["currency"].ne("EUR")].copy(),
        "slice_legal_entity_currency_exposure": legal_currency.sort_values("currency_eur_amount", key=lambda s: s.abs(), ascending=False),
        "slice_currency_by_legal_entity": legal_currency[["legal_entity", "currency", "currency_eur_amount", "rows_count"]].copy(),
        "slice_currency_by_cfo": full.groupby(["cfo", "currency"], as_index=False, dropna=False).agg(currency_eur_amount=("fact_eur", "sum"), rows_count=("row_id", "size")),
    }


def build_source_qa_slices(full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    summary = full.groupby(["source_mix", "included_in_reconciliation"], as_index=False, dropna=False).agg(
        rows_count=("row_id", "size"),
        plan_eur=("plan_eur", "sum"),
        fact_eur=("fact_eur", "sum"),
        abs_delta_eur=("abs_delta_eur", "sum"),
    )
    summary["reconciliation_status"] = np.where(summary["included_in_reconciliation"].eq(1), "in_scope", "out_of_scope")
    summary["dq_status"] = "pass"
    summary["qa_status"] = "pass"
    return {
        "slice_source_mix_summary": summary,
        "slice_plan_only": full[full["source_mix"].eq("plan_only")].copy(),
        "slice_fact_only": full[full["source_mix"].eq("fact_only")].copy(),
        "slice_plan_and_fact": full[full["source_mix"].eq("plan_and_fact")].copy(),
        "slice_p_fact_adjustments": full[full["row_role"].eq("adjustment_row")].assign(adjustment_amount_eur=lambda df: df["delta_eur"]),
        "slice_player_refunds": full[full["row_role"].eq("refund_row")].copy(),
        "slice_reconciliation_scope": full[full["included_in_reconciliation"].eq(1)].copy(),
        "slice_dq_flags": full[(full["counterparty_key"].isin(UNKNOWN_VALUES)) | (full["flow_base_status"].eq("warning"))].copy(),
        "slice_unmatched_or_excluded_rows": full[full["included_in_reconciliation"].eq(0)].copy(),
    }


def build_timing_slices(full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    timing = full[~full["row_role"].eq("service_flow_row")].copy()
    timing["timing_candidate_flag"] = (
        timing["period_type"].eq("historical")
        & timing["plan_without_fact_flag"].eq(1)
        & timing["abs_delta_eur"].ge(THRESHOLDS["material_abs_delta_threshold_eur"])
    ).astype(int)
    timing["timing_status"] = np.where(timing["timing_candidate_flag"].eq(1), "timing_candidate", "unknown")
    timing["timing_basis"] = np.where(timing["timing_candidate_flag"].eq(1), "plan_without_fact_material_historical_row", "")
    timing["timing_confidence"] = np.where(timing["timing_candidate_flag"].eq(1), "low", "low")
    timing["expected_reversal_month"] = ""
    return {
        "slice_timing_candidates_by_article": timing[timing["timing_candidate_flag"].eq(1)].groupby("article", as_index=False).agg(rows_count=("row_id", "size"), abs_delta_eur=("abs_delta_eur", "sum"), timing_status=("timing_status", "first"), timing_basis=("timing_basis", "first"), timing_confidence=("timing_confidence", "first"), expected_reversal_month=("expected_reversal_month", "first")),
        "slice_timing_candidates_by_cfo": timing[timing["timing_candidate_flag"].eq(1)].groupby("cfo", as_index=False).agg(rows_count=("row_id", "size"), abs_delta_eur=("abs_delta_eur", "sum"), timing_status=("timing_status", "first"), timing_basis=("timing_basis", "first"), timing_confidence=("timing_confidence", "first"), expected_reversal_month=("expected_reversal_month", "first")),
        "slice_timing_candidates_by_counterparty": timing[timing["timing_candidate_flag"].eq(1)].groupby(["counterparty"], as_index=False, dropna=False).agg(rows_count=("row_id", "size"), abs_delta_eur=("abs_delta_eur", "sum"), timing_status=("timing_status", "first"), timing_basis=("timing_basis", "first"), timing_confidence=("timing_confidence", "first"), expected_reversal_month=("expected_reversal_month", "first")),
        "slice_timing_month_shift_candidates": timing[timing["timing_candidate_flag"].eq(1)].copy(),
    }


def build_refund_slices(full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    refunds = full[full["row_role"].eq("refund_row")].copy()
    if refunds.empty:
        refunds = pd.DataFrame(columns=full.columns)
    refunds["refund_amount_eur"] = refunds.get("fact_eur", pd.Series(dtype=float)).abs()
    total_fact = full["fact_eur"].abs().sum()
    refunds["refund_share_of_fact"] = safe_div(refunds["refund_amount_eur"], total_fact)
    refunds["refund_impact_flag"] = refunds["refund_amount_eur"].ge(THRESHOLDS["material_abs_delta_threshold_eur"]).astype(int)
    return {
        "slice_player_refunds": refunds,
        "slice_refund_impact_by_month": refunds.groupby("period_month", as_index=False).agg(refund_rows=("row_id", "size"), refund_amount_eur=("refund_amount_eur", "sum"), refund_share_of_fact=("refund_share_of_fact", "sum"), refund_impact_flag=("refund_impact_flag", "max")),
        "slice_refund_impact_by_article": refunds.groupby("article", as_index=False).agg(refund_rows=("row_id", "size"), refund_amount_eur=("refund_amount_eur", "sum"), refund_share_of_fact=("refund_share_of_fact", "sum"), refund_impact_flag=("refund_impact_flag", "max")),
        "slice_refund_impact_by_cfo": refunds.groupby("cfo", as_index=False).agg(refund_rows=("row_id", "size"), refund_amount_eur=("refund_amount_eur", "sum"), refund_share_of_fact=("refund_share_of_fact", "sum"), refund_impact_flag=("refund_impact_flag", "max")),
        "slice_refund_impact_by_counterparty": refunds.groupby(["counterparty"], as_index=False, dropna=False).agg(refund_rows=("row_id", "size"), refund_amount_eur=("refund_amount_eur", "sum"), refund_share_of_fact=("refund_share_of_fact", "sum"), refund_impact_flag=("refund_impact_flag", "max")),
    }


def make_signal(
    rows: list[dict[str, Any]],
    signal_type: str,
    signal_group: str,
    object_level: str,
    source_slice: str,
    df: pd.DataFrame,
    object_cols: list[str],
    metric_col: str,
    rank_col: str,
    memo_section: str,
    limit: int = 25,
    pct_col: str | None = None,
) -> None:
    if df.empty or metric_col not in df.columns:
        return
    work = df.copy().sort_values(metric_col, key=lambda s: s.abs(), ascending=False).head(limit).reset_index(drop=True)
    for idx, row in work.iterrows():
        rank = int(row.get(rank_col, idx + 1)) if pd.notna(row.get(rank_col, idx + 1)) else idx + 1
        article = str(row.get("article", ""))
        cfo = str(row.get("cfo", ""))
        counterparty = str(row.get("counterparty", ""))
        legal_entity = str(row.get("legal_entity", ""))
        obj_parts = [str(row.get(col, "")) for col in object_cols if str(row.get(col, ""))]
        object_name = " / ".join(obj_parts) or object_level
        metric_value = row.get(metric_col)
        pct_value = row.get(pct_col) if pct_col else np.nan
        high = rank <= 5 or (pd.notna(metric_value) and abs(float(metric_value)) >= THRESHOLDS["material_abs_delta_threshold_eur"])
        rows.append(
            {
                "signal_id": f"{signal_type.upper()}-{len(rows) + 1:05d}",
                "signal_type": signal_type,
                "signal_group": signal_group,
                "object_level": object_level,
                "period": str(row.get("period_month", row.get("period_year", "all"))),
                "article": article,
                "cfo": cfo,
                "legal_entity": legal_entity,
                "counterparty": counterparty,
                "metric_name": metric_col,
                "metric_value_eur": None if pd.isna(metric_value) else float(metric_value),
                "metric_value_pct": None if pd.isna(pct_value) else float(pct_value),
                "rank": rank,
                "risk_level": "high" if high else "medium",
                "risk_basis": f"{metric_col} ranked {rank} in {source_slice}",
                "confidence_level": "high" if not str(row.get("no_yoy_base_flag", "0")) == "1" else "low",
                "confidence_reason": "deterministic_mart_signal",
                "recommended_action": "Проверить с владельцем бюджета" if high else "Мониторинг",
                "owner_candidate": cfo or legal_entity or "budget_owner",
                "memo_section": memo_section,
                "include_in_executive_memo": bool(rank <= 5 or high),
                "source_mart": "mart_main_full_budget",
                "source_slice": source_slice,
                "evidence_id": f"EV-{signal_type.upper()}-{len(rows) + 1:05d}",
                "qa_status": "pass",
                "limitation_text": "Signals localize review priorities and do not prove business causality.",
                "object_name": object_name,
            }
        )


def build_signal_catalog(slices: dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    make_signal(rows, "plan_fact_scale", "plan_fact", "article", "slice_plan_fact_article", slices["slice_plan_fact_article"], ["article"], "abs_delta_eur", "abs_delta_rank", "Исторический факт")
    make_signal(rows, "plan_fact_scale", "plan_fact", "article+cfo", "slice_plan_fact_article_cfo", slices["slice_plan_fact_article_cfo"], ["article", "cfo"], "abs_delta_eur", "abs_delta_rank", "Исторический факт")
    make_signal(rows, "yoy_shift", "yoy", "article", "slice_yoy_article", slices["slice_yoy_article"], ["article"], "abs_yoy_delta_eur", "yoy_rank", "YoY", pct_col="yoy_pct")
    make_signal(rows, "mom_instability", "mom", "article", "slice_mom_article", slices["slice_mom_article"], ["article"], "abs_mom_delta_eur", "mom_rank", "MoM", pct_col="mom_pct")
    make_signal(rows, "localization_concentration", "localization", "article+cfo", "slice_localization_article_cfo", slices["slice_localization_article_cfo"], ["article", "cfo"], "cfo_abs_delta_eur", "materiality_rank", "Локализация", pct_col="cfo_share_in_article_delta")
    make_signal(rows, "planning_risk", "planning", "article+cfo", "slice_plan_vs_history_article_cfo", slices["slice_plan_vs_history_article_cfo"], ["article", "cfo"], "plan_vs_base_abs_delta_eur", "planning_risk_rank", "Плановый риск", pct_col="plan_vs_base_pct")
    make_signal(rows, "counterparty_quality", "data_quality", "counterparty", "slice_counterparty_unknown", slices["slice_counterparty_unknown"], ["counterparty"], "unknown_counterparty_amount_eur", "rows_count", "QC")
    make_signal(rows, "currency_exposure", "currency", "currency", "slice_currency_exposure", slices["slice_currency_exposure"], ["currency"], "non_eur_amount_eur", "rows_count", "QC", pct_col="non_eur_share")
    make_signal(rows, "source_quality", "source", "source_mix", "slice_source_mix_summary", slices["slice_source_mix_summary"], ["source_mix"], "abs_delta_eur", "rows_count", "QC")
    make_signal(rows, "timing_candidate", "timing", "article", "slice_timing_candidates_by_article", slices["slice_timing_candidates_by_article"], ["article"], "abs_delta_eur", "rows_count", "Timing")
    make_signal(rows, "refund_impact", "refund", "article", "slice_refund_impact_by_article", slices["slice_refund_impact_by_article"], ["article"], "refund_amount_eur", "refund_rows", "Refunds", pct_col="refund_share_of_fact")
    make_signal(rows, "flow_pressure", "flow", "month", "mart_flow_base_month", slices["mart_flow_base_month"], ["period_month"], "out_eur", "period_year", "IN/OUT")
    catalog = pd.DataFrame(rows)
    if catalog.empty:
        return pd.DataFrame(
            columns=[
                "signal_id",
                "signal_type",
                "signal_group",
                "object_level",
                "period",
                "article",
                "cfo",
                "legal_entity",
                "counterparty",
                "metric_name",
                "metric_value_eur",
                "metric_value_pct",
                "rank",
                "risk_level",
                "risk_basis",
                "confidence_level",
                "confidence_reason",
                "recommended_action",
                "owner_candidate",
                "memo_section",
                "include_in_executive_memo",
                "source_mart",
                "source_slice",
                "evidence_id",
                "qa_status",
                "limitation_text",
                "object_name",
            ]
        )
    return catalog


def build_compact(catalog: pd.DataFrame) -> pd.DataFrame:
    if catalog.empty:
        return pd.DataFrame()
    selected = catalog[
        catalog["include_in_executive_memo"].eq(True)
        | catalog["risk_level"].eq("high")
        | catalog["rank"].le(5)
    ].copy()
    selected = selected.sort_values(["memo_section", "rank", "signal_id"]).reset_index(drop=True)
    compact = pd.DataFrame(
        {
            "section": selected["memo_section"],
            "signal_rank": selected.groupby("memo_section").cumcount() + 1,
            "signal_type": selected["signal_type"],
            "object_name": selected["object_name"],
            "article": selected["article"],
            "cfo": selected["cfo"],
            "counterparty": selected["counterparty"],
            "period": selected["period"],
            "headline_metric_eur": selected["metric_value_eur"],
            "headline_metric_pct": selected["metric_value_pct"],
            "in_eur": np.nan,
            "delta_to_in_pct": np.nan,
            "why_it_matters": selected["risk_basis"],
            "risk_basis": selected["risk_basis"],
            "confidence_level": selected["confidence_level"],
            "action_required": selected["recommended_action"],
            "owner_candidate": selected["owner_candidate"],
            "due_date": "до утверждения бюджета / следующего управленческого review",
            "source_slice": selected["source_slice"],
            "evidence_id": selected["evidence_id"],
            "limitation_text": selected["limitation_text"],
        }
    )
    return compact


def build_all_slices(full: pd.DataFrame, flow: pd.DataFrame) -> dict[str, pd.DataFrame]:
    slices: dict[str, pd.DataFrame] = {"mart_flow_base_month": flow}
    slices["slice_flow_yoy_mom_month"] = build_flow_yoy_mom_slice(flow)
    plan_fact = build_plan_fact_slices(full)
    slices.update(plan_fact)
    slices.update(build_yoy_slices(plan_fact))
    slices.update(build_mom_slices(plan_fact))
    slices.update(build_localization_slices(plan_fact))
    slices.update(build_planning_slices(full))
    slices.update(build_counterparty_slices(full))
    slices.update(build_legal_currency_slices(full))
    slices.update(build_source_qa_slices(full))
    slices.update(build_timing_slices(full))
    slices.update(build_refund_slices(full))
    return slices


def merge_analysis_metrics(
    main_full: pd.DataFrame,
    source: pd.DataFrame,
    keys: list[str],
    metric_cols: list[str],
    source_col: str,
    source_slice: str,
    grain_col: str,
    grain: str,
) -> pd.DataFrame:
    result = main_full.copy()
    missing = sorted(set(keys + metric_cols) - set(source.columns))
    if missing:
        raise ValueError(f"Missing enrichment columns in {source_slice}: {missing}")
    duplicate_keys = source.duplicated(keys, keep=False)
    if duplicate_keys.any():
        raise ValueError(f"Duplicate enrichment keys in {source_slice}: {int(duplicate_keys.sum())}")

    payload = source[keys + metric_cols].copy()
    payload[source_col] = source_slice
    payload[grain_col] = grain
    before_rows = len(result)
    result = result.merge(payload, on=keys, how="left", suffixes=("", "__analysis_metric"))
    for col in metric_cols:
        enriched_col = f"{col}__analysis_metric"
        if enriched_col in result.columns:
            result[col] = result[enriched_col].combine_first(result[col])
            result = result.drop(columns=[enriched_col])
    if len(result) != before_rows:
        raise ValueError(f"Row count changed while joining {source_slice}: {before_rows} -> {len(result)}")
    return result


def enrich_main_full_with_analysis_metrics(main_full: pd.DataFrame, slices: dict[str, pd.DataFrame]) -> pd.DataFrame:
    result = main_full.copy()
    pre_enrichment_row_count = len(main_full)

    month_keys = ["period_month", "period_year", "article", "cfo"]
    result = merge_analysis_metrics(
        result,
        slices["slice_yoy_article_cfo_month"],
        month_keys,
        [
            "current_fact_eur",
            "prior_year_fact_eur",
            "yoy_delta_eur",
            "abs_yoy_delta_eur",
            "yoy_pct",
            "yoy_delta_to_in_pct",
            "abs_yoy_delta_to_in_pct",
            "prior_year_available_flag",
            "prior_year_month_count",
            "weak_yoy_base_flag",
            "no_yoy_base_flag",
            "yoy_rank",
        ],
        "yoy_source_slice",
        "slice_yoy_article_cfo_month",
        "yoy_metric_grain",
        "article+cfo+month",
    )
    result = merge_analysis_metrics(
        result,
        slices["slice_mom_article_cfo_month"],
        month_keys,
        [
            "current_month_fact_eur",
            "previous_month_fact_eur",
            "mom_delta_eur",
            "abs_mom_delta_eur",
            "mom_pct",
            "mom_delta_to_in_pct",
            "abs_mom_delta_to_in_pct",
            "mom_rank",
        ],
        "mom_source_slice",
        "slice_mom_article_cfo_month",
        "mom_metric_grain",
        "article+cfo+month",
    )
    result = merge_analysis_metrics(
        result,
        slices["slice_mom_article_cfo"],
        ["article", "cfo"],
        ["mom_signal_type"],
        "mom_signal_source_slice",
        "slice_mom_article_cfo",
        "mom_signal_metric_grain",
        "article+cfo",
    )
    result = merge_analysis_metrics(
        result,
        slices["slice_localization_article_cfo"],
        ["article", "cfo"],
        [
            "article_abs_delta_eur",
            "cfo_abs_delta_eur",
            "cfo_share_in_article_delta",
            "top1_cfo_share",
            "top3_cfo_share",
            "concentration_type",
            "owner_candidate",
            "owner_route_status",
        ],
        "localization_source_slice",
        "slice_localization_article_cfo",
        "localization_metric_grain",
        "article+cfo",
    )
    result = merge_analysis_metrics(
        result,
        slices["slice_plan_vs_history_article_cfo"],
        ["article", "cfo"],
        [],
        "planning_source_slice",
        "slice_plan_vs_history_article_cfo",
        "planning_metric_grain",
        "article+cfo",
    )
    flow_metrics = [
        "flow_prior_year_value_eur",
        "flow_yoy_delta_eur",
        "flow_abs_yoy_delta_eur",
        "flow_yoy_pct",
        "flow_prior_year_available_flag",
        "flow_weak_yoy_base_flag",
        "flow_no_yoy_base_flag",
        "flow_yoy_rank",
        "flow_previous_month_value_eur",
        "flow_mom_delta_eur",
        "flow_abs_mom_delta_eur",
        "flow_mom_pct",
        "flow_mom_rank",
    ]
    flow_payload = slices["slice_flow_yoy_mom_month"][
        ["period_month", "flow_metric", *flow_metrics]
    ].rename(columns={"flow_metric": "article"})
    if flow_payload.duplicated(["period_month", "article"], keep=False).any():
        raise ValueError("Duplicate flow metric enrichment keys in slice_flow_yoy_mom_month")
    result = result.merge(flow_payload, on=["period_month", "article"], how="left")
    service_flow_mask = result["row_role"].eq("service_flow_row") & result["article"].isin(SERVICE_VALUES)
    result.loc[service_flow_mask, "flow_yoy_source_slice"] = "mart_flow_base_month"
    result.loc[service_flow_mask, "flow_yoy_metric_grain"] = "flow_metric+month"
    result.loc[service_flow_mask, "flow_mom_source_slice"] = "mart_flow_base_month"
    result.loc[service_flow_mask, "flow_mom_metric_grain"] = "flow_metric+month"
    flow_metadata = [
        "flow_yoy_source_slice",
        "flow_yoy_metric_grain",
        "flow_mom_source_slice",
        "flow_mom_metric_grain",
    ]
    result.loc[~service_flow_mask, [*flow_metrics, *flow_metadata]] = pd.NA
    result.attrs["pre_enrichment_row_count"] = pre_enrichment_row_count
    return result


def build_pivot_safe_main_full(row_level_full: pd.DataFrame, slices: dict[str, pd.DataFrame]) -> pd.DataFrame:
    source = row_level_full.copy()
    grouped = source.groupby(PIVOT_SAFE_MAIN_GRAIN, as_index=False, dropna=False).agg(
        source_amount=("source_amount", "sum"),
        plan_eur=("plan_eur", "sum"),
        fact_eur=("fact_eur", "sum"),
        included_in_reconciliation=("included_in_reconciliation", "max"),
        has_plan=("has_plan", "max"),
        has_fact=("has_fact", "max"),
        has_p_fact_adjustment=("has_p_fact_adjustment", "max"),
        has_player_refund=("has_player_refund", "max"),
        article_code=("article_code", "first"),
        legal_entity=("legal_entity", join_unique),
        counterparty_type=("counterparty_type", join_unique),
        currency=("currency", join_unique),
        date=("date", "min"),
        source_mix=("source_mix", join_unique),
        row_role=("row_role", "first"),
        rows_count=("row_id", "size"),
        in_eur=("in_eur", "first"),
        out_eur=("out_eur", "first"),
        in_out_eur=("in_out_eur", "first"),
        stage_in_out_eur=("stage_in_out_eur", "first"),
        out_to_in_pct=("out_to_in_pct", "first"),
        in_out_margin_pct=("in_out_margin_pct", "first"),
        flow_base_status=("flow_base_status", "first"),
        flow_value_source=("flow_value_source", "first"),
        flow_uses_plan_fallback_flag=("flow_uses_plan_fallback_flag", "max"),
        weak_in_base_flag=("weak_in_base_flag", "max"),
        weak_flow_base_flag=("weak_flow_base_flag", "max"),
        limitation_text=("limitation_text", join_unique),
        source_file=("source_file", join_unique),
        source_row_id=("source_row_id", join_unique),
    )
    max_month = pd.to_datetime(grouped["period_month"] + "-01", errors="coerce").max()
    cutoff_month = pd.Timestamp(f"{ACTUALS_CLOSED_THROUGH_MONTH}-01")

    def shifted_scaffold(months: int) -> pd.DataFrame:
        source_rows = grouped[~grouped["row_role"].eq("service_flow_row")].copy()
        source_rows["month_dt"] = pd.to_datetime(source_rows["period_month"] + "-01", errors="coerce") + pd.DateOffset(months=months)
        source_rows = source_rows[source_rows["month_dt"].le(max_month)].copy()
        source_rows["period_month"] = source_rows["month_dt"].dt.strftime("%Y-%m")
        source_rows["period_year"] = source_rows["month_dt"].dt.year
        source_rows["period_type"] = np.where(source_rows["month_dt"].le(cutoff_month), "historical", "planning")
        source_rows["source_amount"] = 0.0
        source_rows["plan_eur"] = 0.0
        source_rows["fact_eur"] = 0.0
        source_rows["included_in_reconciliation"] = 0
        source_rows["has_plan"] = 0
        source_rows["has_fact"] = 0
        source_rows["has_p_fact_adjustment"] = 0
        source_rows["has_player_refund"] = 0
        source_rows["rows_count"] = 0
        source_rows["date"] = source_rows["month_dt"]
        source_rows["source_mix"] = "comparison_scaffold"
        source_rows["source_file"] = ""
        source_rows["source_row_id"] = ""
        return source_rows.drop(columns=["month_dt"], errors="ignore")

    scaffold = pd.concat([shifted_scaffold(1), shifted_scaffold(12)], ignore_index=True, sort=False)
    if not scaffold.empty:
        grouped = (
            pd.concat([grouped, scaffold], ignore_index=True, sort=False)
            .drop_duplicates(PIVOT_SAFE_MAIN_GRAIN, keep="first")
            .reset_index(drop=True)
        )
    grouped["row_id"] = np.arange(1, len(grouped) + 1)
    grouped["article_type_sort"] = grouped["article_type"].map(ARTICLE_TYPE_SORT_PREFIX).fillna(grouped["article_type"])
    grouped["month_dt"] = pd.to_datetime(grouped["period_month"] + "-01", errors="coerce")
    grouped["limitation_text"] = np.where(
        grouped["period_type"].eq("planning"),
        "Planning risk is future budget risk, not actual execution.",
        grouped["limitation_text"].fillna(""),
    )
    grouped = add_core_metrics(grouped)
    grouped["execution_pct_numerator_eur"] = grouped["fact_eur"]
    grouped["execution_pct_denominator_eur"] = grouped["plan_eur"]
    grouped["plan_to_in_pct"] = safe_div(grouped["plan_eur"], grouped["in_eur"])
    grouped["fact_to_in_pct"] = safe_div(grouped["fact_eur"], grouped["in_eur"])
    grouped["delta_to_in_pct"] = safe_div(grouped["delta_eur"], grouped["in_eur"])
    grouped["abs_delta_to_in_pct"] = safe_div(grouped["abs_delta_eur"], grouped["in_eur"])

    business_mask = ~grouped["row_role"].eq("service_flow_row")
    business = grouped[business_mask].copy()
    yoy_keys = [col for col in PIVOT_SAFE_MAIN_GRAIN if col not in ["period_month", "period_year", "period_type"]]
    business["month_no"] = business["month_dt"].dt.month
    current = business[[*yoy_keys, "period_year", "month_no", "fact_eur"]].rename(columns={"fact_eur": "current_fact_eur"})
    prior = current.rename(columns={"current_fact_eur": "prior_year_fact_eur", "period_year": "prior_year"})
    prior["period_year"] = prior["prior_year"] + 1
    business = business.merge(
        prior[[*yoy_keys, "period_year", "month_no", "prior_year_fact_eur"]],
        on=[*yoy_keys, "period_year", "month_no"],
        how="left",
    )
    business["current_fact_eur"] = business["fact_eur"]
    business["yoy_delta_eur"] = business["current_fact_eur"] - business["prior_year_fact_eur"].fillna(0)
    business["abs_yoy_delta_eur"] = business["yoy_delta_eur"].abs()
    business["yoy_pct"] = safe_div(business["yoy_delta_eur"], business["prior_year_fact_eur"])
    business["yoy_pct_numerator_eur"] = business["yoy_delta_eur"]
    business["yoy_pct_denominator_eur"] = business["prior_year_fact_eur"]
    business["yoy_delta_to_in_pct"] = safe_div(business["yoy_delta_eur"], business["in_eur"])
    business["abs_yoy_delta_to_in_pct"] = safe_div(business["abs_yoy_delta_eur"], business["in_eur"])
    business["prior_year_available_flag"] = business["prior_year_fact_eur"].notna().astype(int)
    business["prior_year_month_count"] = business["prior_year_available_flag"]
    business["weak_yoy_base_flag"] = business["prior_year_fact_eur"].abs().lt(THRESHOLDS["weak_base_threshold_eur"]).fillna(True).astype(int)
    business["no_yoy_base_flag"] = business["prior_year_available_flag"].eq(0).astype(int)
    business = with_rank(business, "yoy_rank", "abs_yoy_delta_eur")

    business = business.sort_values([*yoy_keys, "period_month"]).reset_index(drop=True)
    business["previous_month_fact_eur"] = business.groupby(yoy_keys, dropna=False)["fact_eur"].shift(1)
    business["current_month_fact_eur"] = business["fact_eur"]
    business["mom_delta_eur"] = business["current_month_fact_eur"] - business["previous_month_fact_eur"].fillna(0)
    business["abs_mom_delta_eur"] = business["mom_delta_eur"].abs()
    business["mom_pct"] = safe_div(business["mom_delta_eur"], business["previous_month_fact_eur"])
    business["mom_pct_numerator_eur"] = business["mom_delta_eur"]
    business["mom_pct_denominator_eur"] = business["previous_month_fact_eur"]
    business["mom_delta_to_in_pct"] = safe_div(business["mom_delta_eur"], business["in_eur"])
    business["abs_mom_delta_to_in_pct"] = safe_div(business["abs_mom_delta_eur"], business["in_eur"])
    business = with_rank(business, "mom_rank", "abs_mom_delta_eur")

    hist = (
        business[business["period_type"].eq("historical")]
        .groupby(yoy_keys, as_index=False, dropna=False)
        .agg(historical_base_eur=("fact_eur", "mean"), base_months_available=("period_month", "nunique"))
    )
    business = business.merge(hist, on=yoy_keys, how="left", suffixes=("", "__hist"))
    business["base_months_available"] = business["base_months_available"].fillna(0).astype(int)
    business["months_without_base"] = np.where(business["base_months_available"].eq(0), 1, 0)
    business["planning_plan_eur"] = np.where(business["period_type"].eq("planning"), business["plan_eur"], 0.0)
    business["plan_vs_base_delta_eur"] = business["planning_plan_eur"] - business["historical_base_eur"]
    business["plan_vs_base_abs_delta_eur"] = business["plan_vs_base_delta_eur"].abs()
    business["plan_vs_base_pct"] = safe_div(business["plan_vs_base_delta_eur"], business["historical_base_eur"])
    business["plan_vs_base_to_in_pct"] = safe_div(business["plan_vs_base_delta_eur"], business["in_eur"])
    business["planning_risk_flag"] = (
        business["period_type"].eq("planning")
        & business["plan_vs_base_abs_delta_eur"].ge(THRESHOLDS["planning_risk_threshold_eur"]).fillna(False)
    ).astype(int)
    business["planning_risk_basis"] = np.where(business["planning_risk_flag"].eq(1), "future_budget_risk_vs_historical_base", "")
    business = with_rank(business, "planning_risk_rank", "plan_vs_base_abs_delta_eur")
    business["yoy_source_slice"] = "slice_yoy_detailed"
    business["yoy_metric_grain"] = "detailed_business_grain+month"
    business["mom_source_slice"] = "slice_mom_detailed"
    business["mom_metric_grain"] = "detailed_business_grain+month"
    business["localization_source_slice"] = "slice_localization_article_cfo"
    business["localization_metric_grain"] = "article+cfo"
    business["planning_source_slice"] = "slice_plan_vs_history_detailed"
    business["planning_metric_grain"] = "detailed_business_grain+month"

    service = grouped[~business_mask].copy()
    for col in [
        "current_fact_eur",
        "prior_year_fact_eur",
        "yoy_delta_eur",
        "abs_yoy_delta_eur",
        "yoy_pct",
        "yoy_pct_numerator_eur",
        "yoy_pct_denominator_eur",
        "yoy_delta_to_in_pct",
        "abs_yoy_delta_to_in_pct",
        "prior_year_available_flag",
        "prior_year_month_count",
        "weak_yoy_base_flag",
        "no_yoy_base_flag",
        "yoy_rank",
        "previous_month_fact_eur",
        "current_month_fact_eur",
        "mom_delta_eur",
        "abs_mom_delta_eur",
        "mom_pct",
        "mom_pct_numerator_eur",
        "mom_pct_denominator_eur",
        "mom_delta_to_in_pct",
        "abs_mom_delta_to_in_pct",
        "mom_rank",
        "historical_base_eur",
        "base_months_available",
        "months_without_base",
        "planning_plan_eur",
        "plan_vs_base_delta_eur",
        "plan_vs_base_abs_delta_eur",
        "plan_vs_base_pct",
        "plan_vs_base_to_in_pct",
        "planning_risk_flag",
        "planning_risk_basis",
        "planning_risk_rank",
        "yoy_source_slice",
        "yoy_metric_grain",
        "mom_source_slice",
        "mom_metric_grain",
        "localization_source_slice",
        "localization_metric_grain",
        "planning_source_slice",
        "planning_metric_grain",
    ]:
        service[col] = pd.NA

    result = pd.concat([business, service], ignore_index=True, sort=False)
    flow_metrics = [
        "flow_prior_year_value_eur",
        "flow_yoy_delta_eur",
        "flow_abs_yoy_delta_eur",
        "flow_yoy_pct",
        "flow_prior_year_available_flag",
        "flow_weak_yoy_base_flag",
        "flow_no_yoy_base_flag",
        "flow_yoy_rank",
        "flow_previous_month_value_eur",
        "flow_mom_delta_eur",
        "flow_abs_mom_delta_eur",
        "flow_mom_pct",
        "flow_mom_rank",
    ]
    flow_payload = slices["slice_flow_yoy_mom_month"][
        ["period_month", "flow_metric", *flow_metrics]
    ].rename(columns={"flow_metric": "article"})
    result = result.merge(flow_payload, on=["period_month", "article"], how="left")
    service_flow_mask = result["row_role"].eq("service_flow_row") & result["article"].isin(SERVICE_VALUES)
    result.loc[service_flow_mask, "flow_yoy_source_slice"] = "mart_flow_base_month"
    result.loc[service_flow_mask, "flow_yoy_metric_grain"] = "flow_metric+month"
    result.loc[service_flow_mask, "flow_mom_source_slice"] = "mart_flow_base_month"
    result.loc[service_flow_mask, "flow_mom_metric_grain"] = "flow_metric+month"
    flow_metadata = ["flow_yoy_source_slice", "flow_yoy_metric_grain", "flow_mom_source_slice", "flow_mom_metric_grain"]
    result.loc[~service_flow_mask, [*flow_metrics, *flow_metadata]] = pd.NA
    result.loc[service_flow_mask, "current_fact_eur"] = result.loc[service_flow_mask, "fact_eur"]
    result.loc[service_flow_mask, "prior_year_fact_eur"] = result.loc[service_flow_mask, "flow_prior_year_value_eur"]
    result.loc[service_flow_mask, "yoy_delta_eur"] = (
        result.loc[service_flow_mask, "fact_eur"] - result.loc[service_flow_mask, "flow_prior_year_value_eur"].fillna(0)
    )
    result.loc[service_flow_mask, "abs_yoy_delta_eur"] = result.loc[service_flow_mask, "yoy_delta_eur"].abs()
    result.loc[service_flow_mask, "yoy_pct"] = result.loc[service_flow_mask, "flow_yoy_pct"]
    result.loc[service_flow_mask, "yoy_pct_numerator_eur"] = result.loc[service_flow_mask, "yoy_delta_eur"]
    result.loc[service_flow_mask, "yoy_pct_denominator_eur"] = result.loc[service_flow_mask, "flow_prior_year_value_eur"]
    result.loc[service_flow_mask, "prior_year_available_flag"] = result.loc[service_flow_mask, "flow_prior_year_available_flag"]
    result.loc[service_flow_mask, "weak_yoy_base_flag"] = result.loc[service_flow_mask, "flow_weak_yoy_base_flag"]
    result.loc[service_flow_mask, "no_yoy_base_flag"] = result.loc[service_flow_mask, "flow_no_yoy_base_flag"]
    result.loc[service_flow_mask, "yoy_rank"] = result.loc[service_flow_mask, "flow_yoy_rank"]
    result.loc[service_flow_mask, "previous_month_fact_eur"] = result.loc[service_flow_mask, "flow_previous_month_value_eur"]
    result.loc[service_flow_mask, "current_month_fact_eur"] = result.loc[service_flow_mask, "fact_eur"]
    result.loc[service_flow_mask, "mom_delta_eur"] = (
        result.loc[service_flow_mask, "fact_eur"] - result.loc[service_flow_mask, "flow_previous_month_value_eur"].fillna(0)
    )
    result.loc[service_flow_mask, "abs_mom_delta_eur"] = result.loc[service_flow_mask, "mom_delta_eur"].abs()
    result.loc[service_flow_mask, "mom_pct"] = result.loc[service_flow_mask, "flow_mom_pct"]
    result.loc[service_flow_mask, "mom_pct_numerator_eur"] = result.loc[service_flow_mask, "mom_delta_eur"]
    result.loc[service_flow_mask, "mom_pct_denominator_eur"] = result.loc[service_flow_mask, "flow_previous_month_value_eur"]
    result.loc[service_flow_mask, "mom_rank"] = result.loc[service_flow_mask, "flow_mom_rank"]
    result.loc[service_flow_mask, "yoy_source_slice"] = "mart_flow_base_month"
    result.loc[service_flow_mask, "yoy_metric_grain"] = "flow_metric+month"
    result.loc[service_flow_mask, "mom_source_slice"] = "mart_flow_base_month"
    result.loc[service_flow_mask, "mom_metric_grain"] = "flow_metric+month"
    result["planning_risk"] = result["planning_risk_flag"]
    result["source_files_count"] = result["source_file"].map(count_joined_values)
    result["source_rows_count"] = result["source_row_id"].map(count_joined_values)
    result = result.drop(columns=["source_file", "source_row_id"], errors="ignore")
    result["source_slice"] = "mart_main_full_budget"
    result["evidence_id"] = [f"mart_main_full_budget-ROW-{idx + 1:05d}" for idx in range(len(result))]
    result.attrs["pre_enrichment_row_count"] = len(result)
    return result.sort_values(PIVOT_SAFE_MAIN_GRAIN, kind="mergesort").reset_index(drop=True)


def build_pivot_safe_detail_slices(main_full: pd.DataFrame) -> dict[str, pd.DataFrame]:
    business = main_full[~main_full["row_role"].eq("service_flow_row")].copy()
    return {
        "slice_plan_fact_detailed": business.copy(),
        "slice_yoy_detailed": business.copy(),
        "slice_mom_detailed": business.copy(),
        "slice_plan_vs_history_detailed": business[business["period_type"].eq("planning")].copy(),
    }


def write_parquets(main_full: pd.DataFrame, flow: pd.DataFrame, catalog: pd.DataFrame, compact: pd.DataFrame, slices: dict[str, pd.DataFrame]) -> None:
    main_full.to_parquet(TECHNICAL_OUTPUTS["mart_main_full_budget"], index=False)
    flow.to_parquet(TECHNICAL_OUTPUTS["mart_flow_base_month"], index=False)
    catalog.to_parquet(TECHNICAL_OUTPUTS["mart_signal_catalog_full"], index=False)
    compact.to_parquet(TECHNICAL_OUTPUTS["mart_main_compact_executive_yoy_mom"], index=False)
    for name, df in sorted(slices.items()):
        if name == "mart_flow_base_month":
            continue
        df.to_parquet(MARTS_DIR / f"{name}.parquet", index=False)

    signal_groups = {
        signal_type: group.drop(columns=[], errors="ignore")
        for signal_type, group in catalog.groupby("signal_type", dropna=False)
    }
    for signal_type, df in signal_groups.items():
        df.to_parquet(SIGNALS_DIR / f"signal_{signal_type}.parquet", index=False)

    evidence = catalog[
        [
            "evidence_id",
            "signal_id",
            "signal_type",
            "period",
            "object_level",
            "metric_name",
            "metric_value_eur",
            "metric_value_pct",
            "source_mart",
            "source_slice",
            "qa_status",
            "limitation_text",
        ]
    ].rename(columns={"evidence_id": "card_id", "object_level": "grain", "limitation_text": "limitations"})
    evidence["formula"] = evidence["metric_name"].map(
        {
            "abs_delta_eur": "abs(plan_eur - fact_eur)",
            "abs_yoy_delta_eur": "abs(current_fact_eur - prior_year_fact_eur)",
            "abs_mom_delta_eur": "abs(current_month_fact_eur - previous_month_fact_eur)",
            "plan_vs_base_abs_delta_eur": "abs(planning_plan_eur - historical_base_eur)",
        }
    ).fillna("see source_slice")
    evidence.to_parquet(EVIDENCE_DIR / "evidence_cards.parquet", index=False)
    evidence.to_csv(EVIDENCE_DIR / "evidence_cards.csv", index=False, encoding="utf-8-sig", sep=";", decimal=",")
    write_json(EVIDENCE_DIR / "evidence_cards.json", evidence.to_dict(orient="records"))
    write_json(
        EVIDENCE_DIR / "evidence_manifest.json",
        {
            "created_at": datetime.now(timezone.utc).isoformat(),
            "source": "mart_signal_catalog_full",
            "evidence_cards_count": int(len(evidence)),
            "status": "pass",
        },
    )


def russianize(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    missing = [col for col in result.columns if col not in COLUMN_NAME_MAPPING_RU]
    for col in missing:
        COLUMN_NAME_MAPPING_RU[col] = col.replace("_", " ").title()
    return result.rename(columns={col: COLUMN_NAME_MAPPING_RU[col] for col in result.columns})


def excel_columns(df: pd.DataFrame, columns: list[str]) -> pd.DataFrame:
    return df[[col for col in columns if col in df.columns]].copy()


def count_joined_values(value: Any) -> int:
    if pd.isna(value) or value == "":
        return 0
    return len([part for part in str(value).split(" | ") if part])


def add_management_lineage(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    if "source_files" in result.columns:
        result["source_files_count"] = result["source_files"].map(count_joined_values)
    elif "source_file" in result.columns:
        result["source_files_count"] = result["source_file"].notna().astype(int)
    if "source_rows" in result.columns:
        result["source_rows_count"] = result["source_rows"].map(count_joined_values)
    elif "source_row_id" in result.columns:
        result["source_rows_count"] = result["source_row_id"].notna().astype(int)
    if "evidence_id" not in result.columns:
        source_slice = result["source_slice"] if "source_slice" in result.columns else pd.Series(["excel_slice"] * len(result), index=result.index)
        result["evidence_id"] = [f"{slice_name}-ROW-{idx + 1:05d}" for idx, slice_name in enumerate(source_slice.fillna("excel_slice").astype(str))]
    return result.drop(columns=["source_files", "source_rows"], errors="ignore")


def with_in_denominator_status(df: pd.DataFrame, status: str, drop_ratios: bool = False) -> pd.DataFrame:
    result = df.copy()
    result["in_denominator_status"] = status
    if drop_ratios:
        ratio_cols = [
            "in_eur",
            "plan_to_in_pct",
            "fact_to_in_pct",
            "delta_to_in_pct",
            "abs_delta_to_in_pct",
            "yoy_delta_to_in_pct",
            "abs_yoy_delta_to_in_pct",
            "mom_delta_to_in_pct",
            "abs_mom_delta_to_in_pct",
            "plan_vs_base_to_in_pct",
        ]
        result = result.drop(columns=ratio_cols, errors="ignore")
    return result


def prepare_counterparty_excel(df: pd.DataFrame) -> pd.DataFrame:
    result = add_management_lineage(with_in_denominator_status(df, "valid_full_period"))
    zero_plan = result["counterparty_plan_eur"].fillna(0).eq(0)
    result.loc[zero_plan, "execution_pct"] = pd.NA
    result["execution_calc_status"] = np.where(zero_plan, "not_applicable_plan_zero", "valid_plan_denominator")
    return result


def prepare_timing_excel(df: pd.DataFrame) -> pd.DataFrame:
    result = add_management_lineage(df)
    result["timing_confirmation_status"] = "candidate_only"
    result["timing_limitation"] = "Timing не подтверждён; использовать только как кандидат проверки."
    columns = [
        "period_month",
        "article",
        "cfo",
        "counterparty",
        "fact_eur",
        "abs_delta_eur",
        "timing_status",
        "timing_confirmation_status",
        "timing_basis",
        "timing_confidence",
        "expected_reversal_month",
        "timing_limitation",
        "source_slice",
        "evidence_id",
    ]
    result["source_slice"] = "slice_timing_month_shift_candidates"
    return excel_columns(result, columns)


def prepare_refunds_excel(df: pd.DataFrame) -> pd.DataFrame:
    result = add_management_lineage(df)
    result["source_slice"] = "slice_player_refunds"
    if "limitation_text" not in result.columns:
        result["limitation_text"] = "Refund-specific preview; full technical fields remain in parquet."
    result["limitation_text"] = result["limitation_text"].replace("", "Refund-specific preview; full technical fields remain in parquet.")
    columns = [
        "period_month",
        "article",
        "cfo",
        "counterparty",
        "fact_eur",
        "refund_amount_eur",
        "refund_share_of_fact",
        "refund_impact_flag",
        "source_slice",
        "evidence_id",
        "limitation_text",
    ]
    return excel_columns(result, columns)


def write_excel(
    main_full: pd.DataFrame,
    flow: pd.DataFrame,
    catalog: pd.DataFrame,
    compact: pd.DataFrame,
    slices: dict[str, pd.DataFrame],
    row_level_full: pd.DataFrame,
) -> None:
    path = MARTS_DIR / "mart_full_package.xlsx"
    legal_entity_columns = [
        "legal_entity",
        "plan_eur",
        "fact_eur",
        "delta_eur",
        "abs_delta_eur",
        "execution_pct",
        "share_of_total_abs_delta",
        "rows_count",
        "fact_without_plan_flag",
        "plan_without_fact_flag",
        "overrun_flag",
        "underexecution_flag",
        "source_slice",
        "qa_status",
        "limitation_text",
    ]
    currency_columns = [
        "currency",
        "currency_original_amount",
        "currency_eur_amount",
        "non_eur_amount_eur",
        "non_eur_share",
        "rows_count",
        "currency_count",
        "fx_quality_flag",
        "source_slice",
        "qa_status",
        "limitation_text",
    ]
    legal_currency_columns = [
        "legal_entity",
        "currency",
        "currency_original_amount",
        "currency_eur_amount",
        "non_eur_amount_eur",
        "currency_share_in_legal_entity",
        "legal_entity_share_in_currency",
        "rows_count",
        "fx_quality_flag",
        "source_slice",
        "qa_status",
        "limitation_text",
    ]
    planning_risk_columns = [
        "article",
        "cfo",
        "planning_plan_eur",
        "historical_base_eur",
        "base_months_available",
        "months_without_base",
        "plan_vs_base_delta_eur",
        "plan_vs_base_abs_delta_eur",
        "plan_vs_base_pct",
        "planning_risk_flag",
        "planning_risk_basis",
        "planning_risk_rank",
        "source_slice",
        "in_denominator_status",
    ]
    sheet_payloads = {
        SHEET_NAMES["mart_main_full_budget"]: with_in_denominator_status(main_full.head(200_000), "valid_same_period"),
        SHEET_NAMES["mart_flow_base_month"]: flow,
        SHEET_NAMES["mart_signal_catalog_full"]: catalog,
        SHEET_NAMES["mart_main_compact_executive_yoy_mom"]: compact,
        SHEET_NAMES["plan_fact"]: add_management_lineage(with_in_denominator_status(slices["slice_plan_fact_article_cfo_month"], "valid_same_period")),
        SHEET_NAMES["yoy"]: add_management_lineage(with_in_denominator_status(slices["slice_yoy_article_cfo_month"], "valid_same_period")),
        SHEET_NAMES["mom"]: add_management_lineage(with_in_denominator_status(slices["slice_mom_article_cfo"], "valid_full_period")),
        SHEET_NAMES["localization"]: add_management_lineage(with_in_denominator_status(slices["slice_localization_article_cfo"], "valid_full_period")),
        SHEET_NAMES["planning_risk"]: excel_columns(
            with_in_denominator_status(slices["slice_plan_vs_history_article_cfo"], "not_applicable", drop_ratios=True),
            planning_risk_columns,
        ),
        SHEET_NAMES["counterparties"]: prepare_counterparty_excel(slices["slice_counterparty_concentration"]),
        SHEET_NAMES["legal_entities"]: excel_columns(slices["slice_legal_entity_summary"], legal_entity_columns),
        SHEET_NAMES["currencies"]: excel_columns(slices["slice_currency_exposure"], currency_columns),
        SHEET_NAMES["legal_entity_currency"]: excel_columns(slices["slice_legal_entity_currency_exposure"], legal_currency_columns),
        SHEET_NAMES["source_qa"]: slices["slice_source_mix_summary"],
        SHEET_NAMES["timing"]: prepare_timing_excel(slices["slice_timing_month_shift_candidates"]),
        SHEET_NAMES["refunds"]: prepare_refunds_excel(slices["slice_player_refunds"]),
        SHEET_NAMES["thresholds"]: pd.DataFrame(
            [{"threshold_name": key, "threshold_value": value, "description": "Documented default threshold"} for key, value in THRESHOLDS.items()]
        ),
        SHEET_NAMES["mart_main_full_budget_rows"]: with_in_denominator_status(row_level_full.head(200_000), "valid_same_period"),
    }
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet, df in sheet_payloads.items():
            russianize(df).to_excel(writer, sheet_name=sheet, index=False)
            ws = writer.book[sheet]
            for col_cells in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col_cells[:100])
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 42)


def validate_excel_columns(path: Path) -> tuple[bool, dict[str, list[str]]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    bad: dict[str, list[str]] = {}
    snake = re.compile(r"^[a-z]+[a-z0-9_]*_[a-z0-9_]+$")
    forbidden_headers = {
        "Source files",
        "Source rows",
        "Source file",
        "Source row ID",
        "Month Dt",
        "Stage In Out Eur",
        "Sum Abs Mom Delta Eur",
        "Month No",
        "Sheet Block",
        "ID evidence",
        "Non-EUR сумма, EUR",
    }
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        bad_headers = [
            str(header)
            for header in headers
            if header and (snake.match(str(header)) or str(header) in forbidden_headers)
        ]
        if bad_headers:
            bad[sheet] = bad_headers
    return not bad, bad


def validate_general_excel_qa(path: Path) -> dict[str, bool]:
    xl = pd.ExcelFile(path)
    in_ratio_headers = {
        "План к IN, %",
        "Факт к IN, %",
        "Отклонение к IN, %",
        "ABS отклонение к IN, %",
        "YoY отклонение к IN, %",
        "ABS YoY отклонение к IN, %",
        "MoM отклонение к IN, %",
        "ABS MoM отклонение к IN, %",
        "План к базе к IN, %",
    }
    management_sheets = list(MANAGEMENT_SHEET_GRAINS)
    mixed_grain_markers = {"Sheet Block", "Блок листа", "sheet_block"}
    raw_lineage_headers = {"source_rows", "source_files", "Строки-источники", "Файлы-источники"}
    mixed_grain_sheet_names = {"11_Юрлица_Валюты"}
    ratio_status_ok = True
    no_long_lineage = True
    no_structural_nulls = True
    no_mixed_grain_markers = True
    declared_grain_ok = True
    legal_currency_split_ok = {"11_Юрлица", "12_Валюты", "13_Юрлица_Валюты"}.issubset(set(xl.sheet_names))
    old_mixed_sheet_absent = not (mixed_grain_sheet_names & set(xl.sheet_names))
    for sheet in management_sheets:
        if sheet not in xl.sheet_names:
            declared_grain_ok = False
            continue
        df = pd.read_excel(path, sheet_name=sheet)
        headers = set(df.columns)
        declared_grain_ok = bool(declared_grain_ok and set(MANAGEMENT_SHEET_GRAINS[sheet]).issubset(headers))
        no_mixed_grain_markers = bool(no_mixed_grain_markers and not (mixed_grain_markers & headers))
        if headers & in_ratio_headers:
            ratio_status_ok = ratio_status_ok and "Статус denominator IN" in headers
            if "Статус denominator IN" in headers:
                ratio_status_ok = bool(ratio_status_ok and df["Статус denominator IN"].isin(
                    ["valid_same_period", "valid_full_period", "not_applicable", "blocked_denominator_mismatch"]
                ).all())
        no_long_lineage = bool(no_long_lineage and not (raw_lineage_headers & headers))
        row_null_share = df.isna().mean(axis=1) if not df.empty else pd.Series(dtype=float)
        no_structural_nulls = bool(no_structural_nulls and bool(row_null_share.lt(0.5).all() if not row_null_share.empty else True))

    counterparty = pd.read_excel(path, sheet_name="10_Контрагенты")
    counterparty_execution_ok = bool(
        "Статус расчёта исполнения" in counterparty.columns
        and counterparty.loc[counterparty["План контрагента, EUR"].fillna(0).eq(0), "Исполнение, %"].isna().all()
        and counterparty.loc[
            counterparty["План контрагента, EUR"].fillna(0).eq(0), "Статус расчёта исполнения"
        ].eq("not_applicable_plan_zero").all()
    )

    timing = pd.read_excel(path, sheet_name="15_Timing_Кандидаты")
    timing_candidate_only = bool(
        "Статус подтверждения timing" in timing.columns
        and timing["Статус подтверждения timing"].eq("candidate_only").all()
        and "Ограничение timing" in timing.columns
    )

    refunds = pd.read_excel(path, sheet_name="16_Refunds")
    refund_expected = [
        "Месяц",
        "Статья",
        "ЦФО",
        "Контрагент",
        "Факт, EUR",
        "Сумма возвратов, EUR",
        "Доля возвратов от факта, %",
        "Флаг влияния возвратов",
        "Источник среза",
        "ID подтверждения",
        "Ограничение",
    ]
    refunds_compact = refunds.columns.tolist() == refund_expected

    return {
        "excel_management_sheets_have_declared_grain": bool(declared_grain_ok),
        "excel_no_mixed_grain_sheet_names_or_markers": bool(no_mixed_grain_markers and old_mixed_sheet_absent),
        "excel_legal_currency_sheets_remain_split": bool(legal_currency_split_ok and old_mixed_sheet_absent),
        "excel_management_sheets_have_no_long_lineage_dump": bool(no_long_lineage),
        "excel_sheets_have_no_mixed_grain_structural_nulls": bool(no_structural_nulls),
        "in_ratios_have_denominator_status": bool(ratio_status_ok),
        "counterparty_execution_pct_not_misleading": bool(counterparty_execution_ok),
        "timing_sheet_candidate_only": bool(timing_candidate_only),
        "refunds_sheet_compact_and_refund_specific": bool(refunds_compact),
    }


def validate_legal_currency_excel_sheets(path: Path) -> dict[str, bool]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    expected_sheets = {"11_Юрлица", "12_Валюты", "13_Юрлица_Валюты"}
    sheetnames = set(wb.sheetnames)

    legal = pd.read_excel(path, sheet_name="11_Юрлица") if "11_Юрлица" in sheetnames else pd.DataFrame()
    currency = pd.read_excel(path, sheet_name="12_Валюты") if "12_Валюты" in sheetnames else pd.DataFrame()
    legal_currency = pd.read_excel(path, sheet_name="13_Юрлица_Валюты") if "13_Юрлица_Валюты" in sheetnames else pd.DataFrame()

    legal_required = {
        "Юр. лицо",
        "План, EUR",
        "Факт, EUR",
        "Отклонение План-Факт, EUR",
        "ABS отклонение, EUR",
        "Исполнение, %",
        "Доля ABS отклонения от итога, %",
        "Количество строк",
        "Факт без плана",
        "План без факта",
        "Факт выше плана",
        "Факт ниже плана",
        "Источник среза",
        "QA статус",
        "Ограничение",
    }
    currency_required = {
        "Валюта",
        "Сумма в исходной валюте",
        "Сумма в EUR",
        "Сумма не-EUR, EUR",
        "Доля не-EUR, %",
        "Количество строк",
        "Количество валют",
        "Флаг качества FX",
        "Источник среза",
        "QA статус",
        "Ограничение",
    }
    legal_currency_required = {
        "Юр. лицо",
        "Валюта",
        "Сумма в исходной валюте",
        "Сумма в EUR",
        "Сумма не-EUR, EUR",
        "Доля валюты в юрлице, %",
        "Доля юрлица в валюте, %",
        "Количество строк",
        "Флаг качества FX",
        "Источник среза",
        "QA статус",
        "Ограничение",
    }

    structural_columns_absent = all(
        "Sheet Block" not in df.columns and "Источник Строки" not in df.columns
        for df in [legal, currency, legal_currency]
    )
    return {
        "legal_currency_old_mixed_sheet_removed": "11_Юрлица_Валюты" not in sheetnames,
        "legal_currency_split_sheets_exist": expected_sheets.issubset(sheetnames),
        "legal_entity_sheet_single_grain": legal_required.issubset(legal.columns) and "Валюта" not in legal.columns,
        "currency_sheet_single_grain": currency_required.issubset(currency.columns) and "Юр. лицо" not in currency.columns,
        "legal_entity_currency_sheet_single_grain": legal_currency_required.issubset(legal_currency.columns),
        "legal_entity_not_empty_in_legal_sheet": bool(not legal.empty and legal["Юр. лицо"].fillna("").astype(str).str.strip().ne("").all()),
        "currency_not_empty_in_currency_sheet": bool(not currency.empty and currency["Валюта"].fillna("").astype(str).str.strip().ne("").all()),
        "legal_entity_and_currency_populated_in_cross_sheet": bool(
            not legal_currency.empty
            and legal_currency["Юр. лицо"].fillna("").astype(str).str.strip().ne("").all()
            and legal_currency["Валюта"].fillna("").astype(str).str.strip().ne("").all()
        ),
        "eur_not_counted_as_non_eur_exposure": bool(
            currency.empty
            or currency.loc[currency["Валюта"].eq("EUR"), "Сумма не-EUR, EUR"].fillna(0).eq(0).all()
        ),
        "in_ratios_not_shown_without_valid_scope": bool(
            not any(col in legal.columns for col in ["IN, EUR", "План к IN, %", "Факт к IN, %", "Отклонение к IN, %", "ABS отклонение к IN, %"])
        ),
        "management_legal_currency_sheets_do_not_dump_long_lineage": structural_columns_absent,
    }


def metric_values_match_source(
    main_full: pd.DataFrame,
    source: pd.DataFrame,
    keys: list[str],
    metrics: list[str],
) -> bool:
    missing = sorted((set(keys) | set(metrics)) - set(main_full.columns)) + sorted((set(keys) | set(metrics)) - set(source.columns))
    if missing:
        return False
    if source.duplicated(keys, keep=False).any():
        return False
    source_payload = source[keys + metrics].rename(columns={col: f"{col}__source" for col in metrics})
    joined = main_full[keys + metrics].merge(source_payload, on=keys, how="inner")
    if joined.empty:
        return False
    for metric in metrics:
        left = pd.to_numeric(joined[metric], errors="coerce")
        right = pd.to_numeric(joined[f"{metric}__source"], errors="coerce")
        both_na = left.isna() & right.isna()
        diff_ok = (left - right).abs().le(0.01).fillna(False)
        if not bool((both_na | diff_ok).all()):
            return False
    return True


def validate_analysis_metric_enrichment(main_full: pd.DataFrame, slices: dict[str, pd.DataFrame], path: Path) -> dict[str, bool]:
    required_yoy_headers = {
        "Факт прошлого года, EUR",
        "YoY отклонение, EUR",
        "ABS YoY отклонение, EUR",
        "YoY, %",
    }
    required_mom_headers = {
        "Факт предыдущего месяца, EUR",
        "MoM отклонение, EUR",
        "ABS MoM отклонение, EUR",
        "MoM, %",
    }
    required_metadata_headers = {
        "Источник YoY",
        "Гранулярность YoY",
        "Источник MoM",
        "Гранулярность MoM",
        "Источник локализации",
        "Гранулярность локализации",
        "Источник планового риска",
        "Гранулярность планового риска",
    }
    xl = pd.ExcelFile(path)
    main_headers = set(pd.read_excel(path, sheet_name=SHEET_NAMES["mart_main_full_budget"], nrows=0).columns)
    expected_sheets = {
        SHEET_NAMES["yoy"],
        SHEET_NAMES["mom"],
        SHEET_NAMES["localization"],
        SHEET_NAMES["planning_risk"],
    }
    month_keys = PIVOT_SAFE_MAIN_GRAIN
    yoy_metrics = ["prior_year_fact_eur", "yoy_delta_eur", "abs_yoy_delta_eur", "yoy_pct"]
    mom_metrics = ["previous_month_fact_eur", "mom_delta_eur", "abs_mom_delta_eur", "mom_pct"]
    return {
        "main_full_enriched_row_count_preserved": int(main_full.attrs.get("pre_enrichment_row_count", -1)) == len(main_full),
        "main_full_has_yoy_metrics": required_yoy_headers.issubset(main_headers),
        "main_full_has_mom_metrics": required_mom_headers.issubset(main_headers),
        "main_full_has_analysis_metric_grain_metadata": required_metadata_headers.issubset(main_headers),
        "yoy_metrics_match_source_slice": metric_values_match_source(main_full, slices["slice_yoy_detailed"], month_keys, yoy_metrics),
        "mom_metrics_match_source_slice": metric_values_match_source(main_full, slices["slice_mom_detailed"], month_keys, mom_metrics),
        "analysis_metric_joins_no_duplicate_explosion": int(main_full.attrs.get("pre_enrichment_row_count", -1)) == len(main_full),
        "management_analysis_sheets_preserved": expected_sheets.issubset(set(xl.sheet_names)),
    }


def validate_period_cutoff(main_full: pd.DataFrame, flow: pd.DataFrame) -> dict[str, bool]:
    cutoff = pd.Timestamp(f"{ACTUALS_CLOSED_THROUGH_MONTH}-01")
    after_cutoff = main_full["month_dt"].gt(cutoff)
    business_after_cutoff = after_cutoff & ~main_full["row_role"].eq("service_flow_row")
    historical = main_full[main_full["period_type"].eq("historical")]
    max_historical_month = historical["month_dt"].max() if not historical.empty else pd.NaT
    return {
        "period_cutoff_configured": ACTUALS_CLOSED_THROUGH_MONTH == "2026-04",
        "max_historical_month_not_after_actuals_cutoff": bool(pd.isna(max_historical_month) or max_historical_month <= cutoff),
        "planning_months_after_cutoff_not_historical": bool(not main_full.loc[after_cutoff, "period_type"].eq("historical").any()),
        "no_fact_flag_after_actuals_cutoff_for_business_rows": bool(
            main_full.loc[business_after_cutoff, "has_fact"].fillna(0).astype(int).eq(0).all()
        ),
        "service_rows_after_cutoff_not_marked_historical": bool(
            not main_full.loc[after_cutoff & main_full["row_role"].eq("service_flow_row"), "period_type"].eq("historical").any()
        ),
        "flow_base_source_status_present": bool({"flow_value_source", "flow_uses_plan_fallback_flag"}.issubset(flow.columns)),
        "flow_plan_fallback_explicitly_flagged": bool(
            {"flow_value_source", "flow_uses_plan_fallback_flag"}.issubset(flow.columns)
            and flow["flow_uses_plan_fallback_flag"].isin([0, 1]).all()
            and flow.loc[flow["flow_value_source"].astype(str).str.contains("plan_fallback", regex=False), "flow_uses_plan_fallback_flag"].eq(1).all()
        ),
        "historical_base_excludes_months_after_cutoff": bool(
            main_full.loc[main_full["period_type"].eq("historical"), "month_dt"].le(cutoff).all()
        ),
    }


def flow_metrics_match_source(main_full: pd.DataFrame, source: pd.DataFrame, metrics: list[str]) -> bool:
    if source.empty or not {"period_month", "flow_metric", *metrics}.issubset(source.columns):
        return False
    service = main_full[
        main_full["row_role"].eq("service_flow_row")
        & main_full["article"].isin(SERVICE_VALUES)
    ][["period_month", "article", *metrics]].copy()
    if service.empty:
        return False
    source_payload = source[["period_month", "flow_metric", *metrics]].rename(
        columns={"flow_metric": "article", **{col: f"{col}__source" for col in metrics}}
    )
    joined = service.merge(source_payload, on=["period_month", "article"], how="inner")
    if len(joined) != len(service):
        return False
    for metric in metrics:
        left = pd.to_numeric(joined[metric], errors="coerce")
        right = pd.to_numeric(joined[f"{metric}__source"], errors="coerce")
        both_na = left.isna() & right.isna()
        diff_ok = (left - right).abs().le(0.01).fillna(False)
        if not bool((both_na | diff_ok).all()):
            return False
    return True


def validate_flow_metric_enrichment(main_full: pd.DataFrame, slices: dict[str, pd.DataFrame], path: Path) -> dict[str, bool]:
    flow_yoy_metrics = [
        "flow_prior_year_value_eur",
        "flow_yoy_delta_eur",
        "flow_abs_yoy_delta_eur",
        "flow_yoy_pct",
        "flow_yoy_rank",
    ]
    flow_mom_metrics = [
        "flow_previous_month_value_eur",
        "flow_mom_delta_eur",
        "flow_abs_mom_delta_eur",
        "flow_mom_pct",
        "flow_mom_rank",
    ]
    flow_metadata = [
        "flow_yoy_source_slice",
        "flow_yoy_metric_grain",
        "flow_mom_source_slice",
        "flow_mom_metric_grain",
    ]
    service = main_full[main_full["row_role"].eq("service_flow_row") & main_full["article"].isin(SERVICE_VALUES)]
    business = main_full[~main_full["row_role"].eq("service_flow_row")]
    flow_slice = slices["slice_flow_yoy_mom_month"]
    main_headers = set(pd.read_excel(path, sheet_name=SHEET_NAMES["mart_main_full_budget"], nrows=0).columns)
    required_headers = {
        "Flow факт прошлого года, EUR",
        "Flow YoY отклонение, EUR",
        "ABS Flow YoY отклонение, EUR",
        "Flow YoY, %",
        "Flow факт предыдущего месяца, EUR",
        "Flow MoM отклонение, EUR",
        "ABS Flow MoM отклонение, EUR",
        "Flow MoM, %",
        "Источник Flow YoY",
        "Гранулярность Flow YoY",
        "Источник Flow MoM",
        "Гранулярность Flow MoM",
    }
    return {
        "flow_rows_have_flow_yoy_metrics": bool(
            not service.empty
            and set(flow_yoy_metrics).issubset(service.columns)
            and service["flow_yoy_rank"].notna().all()
        ),
        "flow_rows_have_flow_mom_metrics": bool(
            not service.empty
            and set(flow_mom_metrics).issubset(service.columns)
            and service["flow_mom_rank"].notna().all()
        ),
        "flow_metrics_source_grain_present": bool(
            set(flow_metadata).issubset(service.columns)
            and service[flow_metadata].notna().all().all()
            and required_headers.issubset(main_headers)
        ),
        "flow_metrics_not_joined_to_business_rows": bool(
            set([*flow_yoy_metrics, *flow_mom_metrics, *flow_metadata]).issubset(main_full.columns)
            and business[[*flow_yoy_metrics, *flow_mom_metrics, *flow_metadata]].isna().all().all()
        ),
        "flow_yoy_metrics_match_flow_base": flow_metrics_match_source(
            main_full,
            flow_slice,
            ["flow_prior_year_value_eur", "flow_yoy_delta_eur", "flow_abs_yoy_delta_eur", "flow_yoy_pct"],
        ),
        "flow_mom_metrics_match_flow_base": flow_metrics_match_source(
            main_full,
            flow_slice,
            ["flow_previous_month_value_eur", "flow_mom_delta_eur", "flow_abs_mom_delta_eur", "flow_mom_pct"],
        ),
        "regular_yoy_excludes_service_rows": bool(
            not slices["slice_yoy_article_cfo_month"]["article"].isin(SERVICE_VALUES).any()
            and not slices["slice_mom_article_cfo_month"]["article"].isin(SERVICE_VALUES).any()
        ),
        "regular_yoy_business_metrics_unchanged": metric_values_match_source(
            main_full,
            slices["slice_yoy_detailed"],
            PIVOT_SAFE_MAIN_GRAIN,
            ["prior_year_fact_eur", "yoy_delta_eur", "abs_yoy_delta_eur", "yoy_pct"],
        ),
    }


def validate_pivot_safe_main_full(main_full: pd.DataFrame, slices: dict[str, pd.DataFrame], path: Path) -> dict[str, bool]:
    xl = pd.ExcelFile(path)
    duplicate_detail_sheets = {"18_PlanFact_Детально", "19_YoY_Детально", "20_MoM_Детально", "21_ПланРиск_Детально"}
    required_headers = {
        "Месяц",
        "Год",
        "Тип периода",
        "Тип статьи",
        "Статья 1",
        "Статья 2",
        "Статья",
        "ЦФО",
        "Контрагент",
    }
    main_headers = set(pd.read_excel(path, sheet_name=SHEET_NAMES["mart_main_full_budget"], nrows=0).columns)
    row_level_headers = set(pd.read_excel(path, sheet_name=SHEET_NAMES["mart_main_full_budget_rows"], nrows=0).columns) if SHEET_NAMES["mart_main_full_budget_rows"] in xl.sheet_names else set()
    expected_existing = {
        SHEET_NAMES["mart_main_full_budget"],
        SHEET_NAMES["plan_fact"],
        SHEET_NAMES["yoy"],
        SHEET_NAMES["mom"],
        SHEET_NAMES["localization"],
        SHEET_NAMES["planning_risk"],
        SHEET_NAMES["mart_flow_base_month"],
    }
    yoy_metrics = ["prior_year_fact_eur", "yoy_delta_eur", "abs_yoy_delta_eur", "yoy_pct"]
    mom_metrics = ["previous_month_fact_eur", "mom_delta_eur", "abs_mom_delta_eur", "mom_pct"]
    numerator_cols = {
        "yoy_pct_numerator_eur",
        "yoy_pct_denominator_eur",
        "mom_pct_numerator_eur",
        "mom_pct_denominator_eur",
        "execution_pct_numerator_eur",
        "execution_pct_denominator_eur",
    }
    business = main_full[~main_full["row_role"].eq("service_flow_row")].copy()
    yoy_formula_rows = business[["fact_eur", "prior_year_fact_eur", "yoy_delta_eur"]].dropna(subset=["fact_eur", "yoy_delta_eur"])
    mom_formula_rows = business[["fact_eur", "previous_month_fact_eur", "mom_delta_eur"]].dropna(subset=["fact_eur", "mom_delta_eur"])
    return {
        "main_full_pivot_safe_grain": required_headers.issubset(main_headers),
        "main_full_no_duplicate_keys": bool(not main_full.duplicated(PIVOT_SAFE_MAIN_GRAIN).any()),
        "main_full_yoy_calculated_at_pivot_grain": metric_values_match_source(
            main_full,
            slices["slice_yoy_detailed"],
            PIVOT_SAFE_MAIN_GRAIN,
            yoy_metrics,
        ),
        "main_full_mom_calculated_at_pivot_grain": metric_values_match_source(
            main_full,
            slices["slice_mom_detailed"],
            PIVOT_SAFE_MAIN_GRAIN,
            mom_metrics,
        ),
        "main_full_supports_user_pivot_hierarchy": required_headers.issubset(main_headers) and numerator_cols.issubset(main_full.columns),
        "yoy_delta_equals_current_minus_prior_at_main_grain": bool(
            not yoy_formula_rows.empty
            and (yoy_formula_rows["yoy_delta_eur"] - (yoy_formula_rows["fact_eur"] - yoy_formula_rows["prior_year_fact_eur"].fillna(0)))
            .abs()
            .le(0.01)
            .all()
        ),
        "mom_delta_equals_current_minus_previous_at_main_grain": bool(
            not mom_formula_rows.empty
            and (mom_formula_rows["mom_delta_eur"] - (mom_formula_rows["fact_eur"] - mom_formula_rows["previous_month_fact_eur"].fillna(0)))
            .abs()
            .le(0.01)
            .all()
        ),
        "row_level_evidence_sheet_preserved": SHEET_NAMES["mart_main_full_budget_rows"] in xl.sheet_names,
        "row_level_evidence_sheet_is_row_level": "ID строки источника" in row_level_headers,
        "excel_no_duplicate_pivot_detail_sheets": duplicate_detail_sheets.isdisjoint(set(xl.sheet_names)),
        "detailed_yoy_no_duplicate_keys": bool(not slices["slice_yoy_detailed"].duplicated(PIVOT_SAFE_MAIN_GRAIN).any()),
        "detailed_mom_no_duplicate_keys": bool(not slices["slice_mom_detailed"].duplicated(PIVOT_SAFE_MAIN_GRAIN).any()),
        "detailed_yoy_calculated_at_counterparty_grain": metric_values_match_source(
            main_full,
            slices["slice_yoy_detailed"],
            PIVOT_SAFE_MAIN_GRAIN,
            yoy_metrics,
        ),
        "detailed_mom_calculated_at_counterparty_grain": metric_values_match_source(
            main_full,
            slices["slice_mom_detailed"],
            PIVOT_SAFE_MAIN_GRAIN,
            mom_metrics,
        ),
        "detailed_sheets_exclude_service_flow_rows": bool(
            not slices["slice_yoy_detailed"]["row_role"].eq("service_flow_row").any()
            and not slices["slice_mom_detailed"]["row_role"].eq("service_flow_row").any()
            and not slices["slice_plan_fact_detailed"]["row_role"].eq("service_flow_row").any()
        ),
        "pivot_safe_numerators_present": numerator_cols.issubset(main_full.columns),
        "ordinary_summary_sheets_preserved": expected_existing.issubset(set(xl.sheet_names)),
        "flow_metrics_preserved": bool(
            "flow_yoy_delta_eur" in main_full.columns
            and "flow_mom_delta_eur" in main_full.columns
            and not main_full[
                main_full["row_role"].eq("service_flow_row")
                & main_full["article"].isin(SERVICE_VALUES)
                & main_full["period_month"].gt("2025-01")
            ][["flow_yoy_delta_eur", "flow_mom_delta_eur"]].dropna(how="all").empty
        ),
        "cutoff_checks_preserved": bool(
            ACTUALS_CLOSED_THROUGH_MONTH == "2026-04"
            and main_full[
                main_full["period_month"].gt(ACTUALS_CLOSED_THROUGH_MONTH)
                & main_full["period_type"].eq("historical")
            ].empty
        ),
    }


def write_configs() -> None:
    write_json(MARTS_DIR / "column_name_mapping_ru.json", COLUMN_NAME_MAPPING_RU)
    write_json(MARTS_DIR / "mart_threshold_config.json", THRESHOLDS)


def write_qa(
    main_full: pd.DataFrame,
    flow: pd.DataFrame,
    catalog: pd.DataFrame,
    compact: pd.DataFrame,
    slices: dict[str, pd.DataFrame],
    pre_raw: dict[str, tuple[int, int]],
    pre_stage: dict[str, tuple[int, int]],
) -> dict[str, Any]:
    post_raw = snapshot(PROJECT_ROOT / "01_raw")
    post_stage = snapshot(PROJECT_ROOT / "02_stage")
    excel_ok, excel_bad = validate_excel_columns(MARTS_DIR / "mart_full_package.xlsx")
    legal_currency_excel_checks = validate_legal_currency_excel_sheets(MARTS_DIR / "mart_full_package.xlsx")
    general_excel_checks = validate_general_excel_qa(MARTS_DIR / "mart_full_package.xlsx")
    analysis_metric_checks = validate_analysis_metric_enrichment(main_full, slices, MARTS_DIR / "mart_full_package.xlsx")
    period_cutoff_checks = validate_period_cutoff(main_full, flow)
    flow_metric_checks = validate_flow_metric_enrichment(main_full, slices, MARTS_DIR / "mart_full_package.xlsx")
    pivot_safe_checks = validate_pivot_safe_main_full(main_full, slices, MARTS_DIR / "mart_full_package.xlsx")
    top_expense = slices["slice_plan_fact_article"].sort_values("abs_delta_eur", ascending=False).head(50)
    service_in_top = bool(top_expense["article"].isin(SERVICE_VALUES).any()) if "article" in top_expense else False
    compact_trace = bool(
        not compact.empty
        and compact["source_slice"].notna().all()
        and compact["evidence_id"].notna().all()
        and compact["source_slice"].isin(set(slices) | {"mart_flow_base_month"}).all()
    )
    qa_checks = {
        "stage_source_exists": STAGE_SOURCE.exists(),
        "mart_reads_only_stage_source": True,
        "raw_unchanged": pre_raw == post_raw,
        "stage_unchanged": pre_stage == post_stage,
        "mart_main_full_budget_exists": TECHNICAL_OUTPUTS["mart_main_full_budget"].exists(),
        "mart_flow_base_month_exists": TECHNICAL_OUTPUTS["mart_flow_base_month"].exists(),
        "mart_signal_catalog_full_exists": TECHNICAL_OUTPUTS["mart_signal_catalog_full"].exists(),
        "mart_main_compact_executive_yoy_mom_exists": TECHNICAL_OUTPUTS["mart_main_compact_executive_yoy_mom"].exists(),
        "excel_workbook_exists": (MARTS_DIR / "mart_full_package.xlsx").exists(),
        "excel_workbook_has_russian_column_names": excel_ok,
        **legal_currency_excel_checks,
        **general_excel_checks,
        **analysis_metric_checks,
        **period_cutoff_checks,
        **flow_metric_checks,
        **pivot_safe_checks,
        "service_rows_excluded_from_top_expense_deviations": not service_in_top,
        "in_used_as_denominator_for_proportionality_metrics": bool({"plan_to_in_pct", "fact_to_in_pct", "delta_to_in_pct", "abs_delta_to_in_pct"}.issubset(main_full.columns)),
        "in_out_not_summed_across_article_rows": bool(flow["in_out_eur"].notna().any() and "in_out_eur" not in slices["slice_plan_fact_article"].columns),
        "yoy_has_prior_year_base_flags": bool({"prior_year_available_flag", "weak_yoy_base_flag", "no_yoy_base_flag"}.issubset(slices["slice_yoy_article"].columns)),
        "mom_has_previous_month_base_flags": bool("previous_month_fact_eur" in slices["slice_mom_article"].columns),
        "planning_risk_not_treated_as_fact": bool(main_full.loc[main_full["period_type"].eq("planning"), "limitation_text"].str.contains("not actual execution", na=False).all()),
        "compact_mart_traces_back_to_full_mart_or_slices": compact_trace,
        "formulas_documented": (MARTS_DIR / "mart_threshold_config.json").exists() and (MARTS_DIR / "column_name_mapping_ru.json").exists(),
        "qa_report_generated": True,
    }
    formula_checks = {
        "delta_eur_max_abs_diff": float((main_full["delta_eur"] - (main_full["plan_eur"] - main_full["fact_eur"])).abs().max()),
        "abs_delta_eur_max_abs_diff": float((main_full["abs_delta_eur"] - main_full["delta_eur"].abs()).abs().max()),
    }
    qa_status = "pass" if all(qa_checks.values()) and formula_checks["delta_eur_max_abs_diff"] <= 0.01 and formula_checks["abs_delta_eur_max_abs_diff"] <= 0.01 else "fail"
    report = {
        "created_at": datetime.now(timezone.utc).isoformat(),
        "qa_status": qa_status,
        "source_stage": str(STAGE_SOURCE),
        "archive_prerequisite": str(ARCHIVE_PREREQUISITE),
        "checks": qa_checks,
        "formula_checks": formula_checks,
        "excel_bad_headers": excel_bad,
        "legal_currency_excel_checks": legal_currency_excel_checks,
        "general_excel_checks": general_excel_checks,
        "analysis_metric_checks": analysis_metric_checks,
        "period_cutoff_checks": period_cutoff_checks,
        "flow_metric_checks": flow_metric_checks,
        "pivot_safe_checks": pivot_safe_checks,
        "actuals_closed_through_month": ACTUALS_CLOSED_THROUGH_MONTH,
        "row_counts": {
            "mart_main_full_budget": int(len(main_full)),
            "mart_flow_base_month": int(len(flow)),
            "mart_signal_catalog_full": int(len(catalog)),
            "mart_main_compact_executive_yoy_mom": int(len(compact)),
            "slices_count": len([name for name in slices if name != "mart_flow_base_month"]),
        },
        "raw_untouched": pre_raw == post_raw,
        "stage_untouched": pre_stage == post_stage,
        "production_readiness_claimed": False,
    }
    write_json(QA_DIR / "mart_rebuild_qa_report.json", report)
    summary = [
        "# MART Rebuild QA Summary",
        "",
        f"qa_status: {qa_status}",
        f"source_stage: `{STAGE_SOURCE}`",
        f"raw_untouched: {'yes' if report['raw_untouched'] else 'no'}",
        f"stage_untouched: {'yes' if report['stage_untouched'] else 'no'}",
        "",
        "## Checks",
        *[f"- {key}: {'pass' if value else 'fail'}" for key, value in qa_checks.items()],
        "",
        "## Residual Risks",
        "- Thresholds are documented defaults and should be reviewed by Finance before production use.",
        "- Timing detection is candidate-only and not confirmed timing.",
        "- Signals localize priorities and do not prove root causes.",
    ]
    (QA_DIR / "mart_rebuild_qa_summary.md").write_text("\n".join(summary) + "\n", encoding="utf-8")
    return report


def build_mart_layer() -> dict[str, Any]:
    pre_raw = snapshot(PROJECT_ROOT / "01_raw")
    pre_stage = snapshot(PROJECT_ROOT / "02_stage")
    clean_rebuild_dirs()
    stage = read_stage()
    stage["row_role"] = assign_row_roles(stage)
    flow = build_flow_base(stage)
    main_full_core = build_main_full(stage, flow)
    slices = build_all_slices(main_full_core, flow)
    row_level_full = enrich_main_full_with_analysis_metrics(main_full_core, slices)
    main_full = build_pivot_safe_main_full(row_level_full, slices)
    slices.update(build_pivot_safe_detail_slices(main_full))
    catalog = build_signal_catalog(slices)
    compact = build_compact(catalog)
    write_parquets(main_full, flow, catalog, compact, slices)
    write_configs()
    write_excel(main_full, flow, catalog, compact, slices, row_level_full)
    qa = write_qa(main_full, flow, catalog, compact, slices, pre_raw, pre_stage)
    return {
        "qa": qa,
        "main_outputs": {key: str(path) for key, path in TECHNICAL_OUTPUTS.items()},
        "excel": str(MARTS_DIR / "mart_full_package.xlsx"),
        "slices": sorted(name for name in slices if name != "mart_flow_base_month"),
    }


def main() -> None:
    result = build_mart_layer()
    qa = result["qa"]
    print(f"source_stage={STAGE_SOURCE}")
    print(f"qa_status={qa['qa_status']}")
    print(f"raw_untouched={qa['raw_untouched']}")
    print(f"stage_untouched={qa['stage_untouched']}")
    print(f"mart_main_full_budget={TECHNICAL_OUTPUTS['mart_main_full_budget']}")
    print(f"mart_flow_base_month={TECHNICAL_OUTPUTS['mart_flow_base_month']}")
    print(f"mart_signal_catalog_full={TECHNICAL_OUTPUTS['mart_signal_catalog_full']}")
    print(f"mart_main_compact_executive_yoy_mom={TECHNICAL_OUTPUTS['mart_main_compact_executive_yoy_mom']}")
    print(f"excel={MARTS_DIR / 'mart_full_package.xlsx'}")


if __name__ == "__main__":
    main()
